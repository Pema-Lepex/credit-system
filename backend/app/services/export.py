"""ExportService -- "download my data", and the archive escape hatch.

TENANCY IS THE WHOLE BALLGAME HERE (read this before touching a query)
----------------------------------------------------------------------
An export is the one operation that takes a large slice of the database and hands
it to a human as a file. A missing ``business_id`` filter anywhere in this module
does not produce a subtly wrong number on a dashboard -- it silently ships another
shop's customer list, phone numbers and debts to a stranger. That is the worst bug
this codebase could have.

So every dataset builder below starts from ``self.scope_id`` (the ONE business the
caller is permitted to touch, decided by BaseService.scope_id -- never from a
client-supplied id), and every ``select`` carries an explicit, commented
``business_id ==`` predicate. There is no shared "base query" helper that a future
dataset could forget to call: the filter is written out, visibly, in each one.

LIFECYCLE
---------
    PENDING -> RUNNING -> READY (file attached, expires_at set)
                       \\-> FAILED (error recorded, no file)
    READY   -> (purged: file AND row deleted outright)

Exported files self-destruct after ``settings.EXPORT_TTL_HOURS``. Nothing generated
is kept forever -- that is what stops a free-tier disk filling with stale
spreadsheets, and it is why ``purge_stale`` exists.

WHY THE ROW GOES TOO
--------------------
An earlier version kept the ExportJob row as a receipt and only flipped it to
EXPIRED. That left the exports table filling up with dead entries the user could not
act on, and it kept a record of *what was exported* (dataset names, row counts,
filters) around indefinitely -- which is exactly the kind of residue a "delete my
export" button is supposed to remove. The receipt now lives in ``audit_log``, where
an append-only trail belongs; the ExportJob row is deleted outright, satisfying the
spec's "all deletion operations must be logged" without keeping the payload.

EXPIRED remains in the enum for historical rows written before this change.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# `select` MUST come from sqlmodel here, not sqlalchemy. They look identical and are
# not: sqlmodel.select(Model) returns a SelectOfScalar, which Session.exec() unwraps
# into model instances. sqlalchemy.select(Model) returns a plain Select, which exec()
# hands back as Row objects -- so `job.state` raises AttributeError and the whole
# exports list 500s ("Could not load your exports"). Every other service imports it
# from sqlmodel; this file was the odd one out.
from sqlmodel import Session, col, select

from app.core.config import settings
from app.core.errors import NotFoundError, ValidationError
from app.core.security import Permission
from app.models.base import utcnow
from app.models.business import Business
from app.models.catalog import Product, Service
from app.models.credit import Credit, CreditItem, Payment
from app.models.customer import Customer
from app.models.enums import (
    AuditAction,
    ExportFormat,
    ExportState,
    FileKind,
    ReportPeriod,
)
from app.models.file import FileAsset
from app.models.retention import ArchiveBatch, AuditLog, ExportJob
from app.models.expense import Expense, ExpenseCategory
from app.services.accounting import UNCATEGORISED, AccountingService
from app.services.base import BaseService, ServiceContext
from app.services.reports import ReportService
from app.storage.service import StorageService
from app.utils.dates import end_of_day, ensure_utc, start_of_day
from app.utils.pagination import Page, PageInput, paginate

# The datasets a caller may ask for. A whitelist, not a getattr() lookup: an export
# is the last place you want a client-supplied string reaching the ORM.
DATASETS: tuple[str, ...] = (
    "customers",
    "credits",
    "payments",
    "products",
    "services",
    "business",
    "reports",
    "expenses",
    "expense_summary",
    "profit_loss",
    "cash_flow",
    "aging_receivable",
    "tax_summary",
)

_CONTENT_TYPES: dict[str, str] = {
    "csv": "text/csv",
    "zip": "application/zip",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "json": "application/json",
    "pdf": "application/pdf",
}


@dataclass(slots=True)
class Dataset:
    """One table's worth of an export: a name, a header row, and the rows."""

    name: str
    headers: list[str]
    rows: list[list[Any]]

    @property
    def row_count(self) -> int:
        return len(self.rows)


@dataclass(frozen=True, slots=True)
class ExportBundle:
    """The finished artefact, before it is handed to storage."""

    data: bytes
    filename: str
    content_type: str
    row_count: int


class ExportService(BaseService):
    # ================================================================== create
    async def create_export(
        self,
        ctx: ServiceContext,
        format: ExportFormat,
        datasets: list[str],
        filters: dict[str, Any] | None = None,
    ) -> ExportJob:
        """Run an export end to end and return the (READY or FAILED) job."""
        self.require(Permission.EXPORT_CREATE)
        business = self.get_business()

        wanted = [d.strip().lower() for d in datasets if d and d.strip()]
        unknown = [d for d in wanted if d not in DATASETS]
        if unknown:
            raise ValidationError(
                f"Unknown dataset(s): {', '.join(unknown)}. "
                f"Choose from: {', '.join(DATASETS)}",
                field="datasets",
            )
        if not wanted:
            raise ValidationError("Select at least one dataset to export", field="datasets")

        job = ExportJob(
            business_id=self.scope_id,  # tenancy boundary: the job belongs to ONE business
            format=format,
            state=ExportState.PENDING,
            datasets=wanted,
            # ExportJob.filters is a JSON column, and json.dumps cannot serialise a
            # date/Decimal -- the INSERT dies with "Object of type date is not JSON
            # serializable". Every dated export (the whole Reports page) hit this.
            # Normalised HERE, at the one place filters are written, so no caller has
            # to remember. Readers already cope: _as_date() parses ISO strings.
            filters=_json_safe(filters or {}),
            requested_by_user_id=ctx.user.id if ctx.user else None,
        )
        self.session.add(job)
        self.session.flush()

        return await self._run(job, business, wanted, filters or {})

    async def _run(
        self,
        job: ExportJob,
        business: Business,
        wanted: list[str],
        filters: dict[str, Any],
    ) -> ExportJob:
        job.state = ExportState.RUNNING
        self.session.add(job)
        self.session.flush()

        try:
            sets = [self._build_dataset(name, filters) for name in wanted]
            bundle = self._render(ExportFormat(job.format), sets, business)

            asset = await StorageService(self.session).upload(
                business_id=self.scope_id,  # tenancy boundary: file is filed under this business
                kind=FileKind.EXPORT,
                filename=bundle.filename,
                data=bundle.data,
                content_type=bundle.content_type,
                user_id=job.requested_by_user_id,
                expires_in_hours=settings.EXPORT_TTL_HOURS,
            )
            # The job is the asset's only referrer. Without this the nightly orphan
            # sweep would delete the file 24h before the TTL even matters.
            StorageService(self.session).attach(asset.id)

            job.file_id = asset.id
            job.row_count = sum(s.row_count for s in sets)
            job.size_bytes = len(bundle.data)
            job.state = ExportState.READY
            job.expires_at = utcnow() + timedelta(hours=settings.EXPORT_TTL_HOURS)
            job.completed_at = utcnow()
            job.error = None

            self.audit(
                AuditAction.EXPORT,
                "export_job",
                job.id,
                f"Exported {', '.join(wanted)} as {ExportFormat(job.format).value} "
                f"({job.row_count} rows, {job.size_bytes} bytes)",
            )
        except Exception as exc:  # noqa: BLE001
            # A failed export must leave a FAILED row the user can see, not vanish.
            # (StorageError from an oversized file lands here too -- MAX_UPLOAD_MB
            # applies to exports as much as to uploads.)
            job.state = ExportState.FAILED
            job.error = str(exc)[:1000]
            job.completed_at = utcnow()
            job.file_id = None
            # A FAILED row expires on the same clock as a successful one. Without
            # this it has no expires_at, so purge_stale never matches it and the
            # error sits in the user's exports table forever.
            job.expires_at = utcnow() + timedelta(hours=settings.EXPORT_TTL_HOURS)
            self.session.add(job)
            self.session.flush()
            return job

        self.session.add(job)
        self.session.flush()
        return job

    # ================================================================== render
    def _render(
        self, format: ExportFormat, sets: list[Dataset], business: Business
    ) -> ExportBundle:
        stamp = f"{utcnow():%Y%m%d-%H%M%S}"
        slug = business.slug
        rows = sum(s.row_count for s in sets)

        if format is ExportFormat.XLSX:
            return ExportBundle(
                data=self._to_xlsx(sets),
                filename=f"{slug}-export-{stamp}.xlsx",
                content_type=_CONTENT_TYPES["xlsx"],
                row_count=rows,
            )

        if format is ExportFormat.JSON:
            return ExportBundle(
                data=self._to_json(sets),
                filename=f"{slug}-export-{stamp}.json",
                content_type=_CONTENT_TYPES["json"],
                row_count=rows,
            )

        if format is ExportFormat.PDF:
            return ExportBundle(
                data=self._to_pdf(sets, business),
                filename=f"{slug}-export-{stamp}.pdf",
                content_type=_CONTENT_TYPES["pdf"],
                row_count=rows,
            )

        # CSV. One dataset -> one .csv. More than one -> a ZIP of one .csv each.
        # A CSV is a SINGLE table: one header row, then homogeneous rows. Customers
        # and payments have different columns, so stacking them into one file would
        # produce something no spreadsheet, no pandas.read_csv and no human can parse
        # correctly. The ZIP keeps each table a valid CSV.
        if len(sets) == 1:
            return ExportBundle(
                data=_csv_bytes(sets[0]),
                filename=f"{slug}-{sets[0].name}-{stamp}.csv",
                content_type=_CONTENT_TYPES["csv"],
                row_count=rows,
            )

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for ds in sets:
                zf.writestr(f"{ds.name}.csv", _csv_bytes(ds))
        return ExportBundle(
            data=buf.getvalue(),
            filename=f"{slug}-export-{stamp}.zip",
            content_type=_CONTENT_TYPES["zip"],
            row_count=rows,
        )

    @staticmethod
    def _to_xlsx(sets: list[Dataset]) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]  # drop the default empty sheet
        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF")

        for ds in sets:
            # Excel sheet names are capped at 31 chars and cannot contain []:*?/\.
            sheet = wb.create_sheet(ds.name[:31])
            sheet.append(ds.headers)
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for row in ds.rows:
                sheet.append([_cell(v) for v in row])
            sheet.freeze_panes = "A2"
            for i, name in enumerate(ds.headers, start=1):
                width = max(
                    [len(str(name))] + [len(str(r[i - 1])) for r in ds.rows[:200] if len(r) >= i]
                )
                sheet.column_dimensions[get_column_letter(i)].width = min(50, max(11, width + 3))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _to_json(sets: list[Dataset]) -> bytes:
        payload = {
            ds.name: [dict(zip(ds.headers, (_cell(v) for v in row), strict=True)) for row in ds.rows]
            for ds in sets
        }
        return json.dumps(
            {"generated_at": utcnow().isoformat(), "data": payload},
            indent=2,
            default=str,
        ).encode("utf-8")

    @staticmethod
    def _to_pdf(sets: list[Dataset], business: Business) -> bytes:
        """A tabular PDF. Deliberately plain -- a data dump is for reading, not for
        handing to a customer; the pretty documents live in ReportService."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm, title=f"{business.name} export",
        )
        styles = getSampleStyleSheet()
        flow: list[Any] = []
        for i, ds in enumerate(sets):
            if i:
                flow.append(PageBreak())
            flow.append(Paragraph(f"{business.name} - {ds.name}", styles["Title"]))
            flow.append(Spacer(1, 4 * mm))
            data = [ds.headers] + [[str(_cell(v) or "") for v in r] for r in ds.rows]
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                    ]
                )
            )
            flow.append(table)
        doc.build(flow)
        return buf.getvalue()

    # ================================================================ datasets
    # EVERY query below is scoped to self.scope_id. See the module docstring.
    def _build_dataset(self, name: str, filters: dict[str, Any]) -> Dataset:
        builders = {
            "customers": self._customers,
            "credits": self._credits,
            "payments": self._payments,
            "products": self._products,
            "services": self._services,
            "business": self._business_dataset,
            "reports": self._reports,
            "expenses": self._expenses,
            "expense_summary": self._expense_summary,
            "profit_loss": self._profit_loss,
            "cash_flow": self._cash_flow,
            "aging_receivable": self._aging_receivable,
            "tax_summary": self._tax_summary,
        }
        return builders[name](filters)

    def _customers(self, filters: dict[str, Any]) -> Dataset:  # noqa: ARG002
        # Customers are never archived individually (only closed credits are), so
        # there is no archive filter here -- a purged batch must not take the
        # customer record with it.
        stmt = (
            select(Customer)
            .where(
                Customer.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Customer.deleted_at).is_(None),
            )
            .order_by(col(Customer.code).asc())
        )
        rows = [
            [
                c.code,
                c.name,
                c.phone,
                c.email,
                c.address,
                c.city,
                c.status.value if hasattr(c.status, "value") else str(c.status),
                c.credit_score,
                _money(c.credit_limit),
                _money(c.total_credit),
                _money(c.total_paid),
                _money(c.outstanding_balance),
                c.credit_count,
                c.overdue_count,
                _dt(c.last_credit_at),
                _dt(c.last_payment_at),
                _dt(c.created_at),
            ]
            for c in self.session.execute(stmt).scalars().all()
        ]
        return Dataset(
            name="customers",
            headers=[
                "Code", "Name", "Phone", "Email", "Address", "City", "Status",
                "Credit score", "Credit limit", "Total credit", "Total paid",
                "Outstanding", "Credits", "Overdue", "Last credit", "Last payment",
                "Created",
            ],
            rows=rows,
        )

    def _credits(self, filters: dict[str, Any]) -> Dataset:
        stmt = (
            select(Credit, Customer.code, Customer.name)
            .join(Customer, col(Credit.customer_id) == col(Customer.id))
            .where(
                Credit.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Credit.deleted_at).is_(None),
            )
            .order_by(col(Credit.issued_date).desc())
        )
        stmt = _apply_date_filter(stmt, col(Credit.issued_date), filters)
        if batch_id := filters.get("archive_batch_id"):
            stmt = stmt.where(Credit.archive_batch_id == batch_id)
        elif not filters.get("include_archived"):
            stmt = stmt.where(col(Credit.archived_at).is_(None))

        rows = [
            [
                c.number,
                code,
                name,
                c.issued_date.isoformat(),
                c.due_date.isoformat(),
                c.status.value if hasattr(c.status, "value") else str(c.status),
                c.currency,
                _money(c.subtotal),
                _money(c.discount_amount),
                _money(c.tax_amount),
                _money(c.grand_total),
                _money(c.amount_paid),
                _money(c.remaining_amount),
                c.notes,
                _dt(c.created_at),
            ]
            for c, code, name in self.session.execute(stmt).all()
        ]
        return Dataset(
            name="credits",
            headers=[
                "Number", "Customer code", "Customer", "Issued", "Due", "Status",
                "Currency", "Subtotal", "Discount", "Tax", "Total", "Paid",
                "Remaining", "Notes", "Created",
            ],
            rows=rows,
        )

    def _payments(self, filters: dict[str, Any]) -> Dataset:
        stmt = (
            select(Payment, Credit.number, Customer.name)
            # OUTER join, and that is load-bearing. An ACCOUNT payment settles the
            # customer's balance and names no credit (credit_id IS NULL), so an inner
            # join silently DROPPED it -- money missing from a financial report, with
            # nothing to indicate anything had been left out. The customer join stays
            # inner: every payment has a customer, and one that did not would be a
            # corruption worth failing on.
            .join(Credit, col(Payment.credit_id) == col(Credit.id), isouter=True)
            .join(Customer, col(Payment.customer_id) == col(Customer.id))
            .where(
                Payment.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Payment.deleted_at).is_(None),
            )
            .order_by(col(Payment.paid_at).desc())
        )
        # paid_at is an INSTANT, so a date filter must be widened to the business's
        # local day bounds in UTC. Comparing a timestamp column directly against a
        # bare date would silently drop everything after 00:00 on the end date.
        tz = self.get_business().timezone
        if start := _as_date(filters.get("start")):
            stmt = stmt.where(col(Payment.paid_at) >= start_of_day(start, tz))
        if end := _as_date(filters.get("end")):
            stmt = stmt.where(col(Payment.paid_at) < end_of_day(end, tz))
        if batch_id := filters.get("archive_batch_id"):
            stmt = stmt.where(Payment.archive_batch_id == batch_id)
        elif not filters.get("include_archived"):
            stmt = stmt.where(col(Payment.archived_at).is_(None))
        if not filters.get("include_voided", True):
            stmt = stmt.where(col(Payment.voided_at).is_(None))

        rows = [
            [
                p.number,
                # NULL for an account payment. Says so, rather than leaving a blank
                # cell that reads like data went missing.
                credit_number or "Account balance",
                customer_name,
                _dt(p.paid_at),
                _money(p.amount),
                p.method.value if hasattr(p.method, "value") else str(p.method),
                p.provider or "",
                p.reference,
                _money(p.balance_after),
                "YES" if p.voided_at else "",
                p.void_reason,
                p.notes,
            ]
            for p, credit_number, customer_name in self.session.execute(stmt).all()
        ]
        return Dataset(
            name="payments",
            headers=[
                "Number", "Credit", "Customer", "Paid at (UTC)", "Amount", "Method",
                "Bank / provider", "Reference", "Balance after", "Voided",
                "Void reason", "Notes",
            ],
            rows=rows,
        )

    def _products(self, filters: dict[str, Any]) -> Dataset:  # noqa: ARG002
        stmt = (
            select(Product)
            .where(
                Product.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Product.deleted_at).is_(None),
            )
            .order_by(col(Product.name).asc())
        )
        rows = [
            [
                p.name, p.sku, p.barcode, _money(p.price), _money(p.cost_price),
                str(p.stock_quantity), p.unit, "YES" if p.is_active else "NO",
                p.description, _dt(p.created_at),
            ]
            for p in self.session.execute(stmt).scalars().all()
        ]
        return Dataset(
            name="products",
            headers=[
                "Name", "SKU", "Barcode", "Price", "Cost", "Stock", "Unit",
                "Active", "Description", "Created",
            ],
            rows=rows,
        )

    def _services(self, filters: dict[str, Any]) -> Dataset:  # noqa: ARG002
        stmt = (
            select(Service)
            .where(
                Service.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Service.deleted_at).is_(None),
            )
            .order_by(col(Service.name).asc())
        )
        rows = [
            [
                s.name, s.code, _money(s.price), s.duration_minutes,
                "YES" if s.is_active else "NO", s.description, _dt(s.created_at),
            ]
            for s in self.session.execute(stmt).scalars().all()
        ]
        return Dataset(
            name="services",
            headers=["Name", "Code", "Price", "Duration (min)", "Active", "Description", "Created"],
            rows=rows,
        )

    # NOT named _business: BaseService caches the tenant on self._business, and a
    # method of the same name would shadow it.
    def _business_dataset(self, filters: dict[str, Any]) -> Dataset:  # noqa: ARG002
        # get_business() already resolves through scope_id -- there is no way to name
        # another tenant here.
        b = self.get_business()
        pairs = [
            ("Name", b.name), ("Slug", b.slug), ("Description", b.description),
            ("Email", b.email), ("Phone", b.phone), ("WhatsApp", b.whatsapp_number),
            ("Website", b.website), ("Address", b.address), ("City", b.city),
            ("Country", b.country), ("Currency", b.currency),
            ("Currency symbol", b.currency_symbol), ("Timezone", b.timezone),
            ("Locale", b.locale), ("Tax %", str(b.tax_percentage)),
            ("Retention policy", str(b.retention_policy)),
            ("Storage quota (MB)", b.storage_quota_mb),
            ("Reminders enabled", "YES" if b.reminders_enabled else "NO"),
            ("Reminder days before", ", ".join(str(d) for d in b.reminder_days_before)),
            ("Created", _dt(b.created_at)),
        ]
        return Dataset(
            name="business",
            headers=["Field", "Value"],
            rows=[[k, v] for k, v in pairs],
        )

    def _reports(self, filters: dict[str, Any]) -> Dataset:
        """The report breakdown rows, as a table. Reuses ReportService so the export
        and the on-screen report can never disagree."""
        period = ReportPeriod(str(filters.get("period", ReportPeriod.MONTHLY.value)).upper())
        report = ReportService(self.ctx).generate(
            period,
            start=_as_date(filters.get("start")),
            end=_as_date(filters.get("end")),
        )
        rows = [
            [
                r.label,
                float(r.credits_issued),
                r.credit_count,
                float(r.collected),
                r.payment_count,
                float(r.net),
            ]
            for r in report.rows
        ]
        return Dataset(
            name="reports",
            headers=["Period", "Credits issued", "Credits", "Collected", "Payments", "Net"],
            rows=rows,
        )

    def _expenses(self, filters: dict[str, Any]) -> Dataset:
        """Every expense line in range. The detail behind the summary below.

        ``expense_date`` is a calendar DATE (see models/expense.py), so unlike
        ``_payments`` it is compared against bare dates -- no timezone widening,
        because there is no instant to widen.
        """
        stmt = (
            select(Expense, ExpenseCategory.name)
            # OUTER: an uncategorised expense is ordinary, and an inner join would
            # silently drop it -- money missing from a financial export.
            .join(
                ExpenseCategory,
                col(Expense.category_id) == col(ExpenseCategory.id),
                isouter=True,
            )
            .where(
                Expense.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Expense.deleted_at).is_(None),
            )
            .order_by(col(Expense.expense_date).desc())
        )
        if start := _as_date(filters.get("start")):
            stmt = stmt.where(col(Expense.expense_date) >= start)
        if end := _as_date(filters.get("end")):
            stmt = stmt.where(col(Expense.expense_date) <= end)

        rows = [
            [
                e.expense_date.isoformat(),
                category_name or UNCATEGORISED,
                e.vendor_name or "",
                _money(e.amount),
                e.payment_method.value
                if hasattr(e.payment_method, "value")
                else str(e.payment_method),
                e.provider or "",
                e.reference,
                e.notes,
            ]
            for e, category_name in self.session.execute(stmt).all()
        ]
        return Dataset(
            name="expenses",
            headers=[
                "Date",
                "Category",
                "Vendor",
                "Amount",
                "Method",
                "Bank / provider",
                "Reference",
                "Notes",
            ],
            rows=rows,
        )

    def _expense_summary(self, filters: dict[str, Any]) -> Dataset:
        """The grouped breakdown. Reuses AccountingService so the export and the
        on-screen report can never disagree -- same reasoning as _reports."""
        report = AccountingService(self.ctx).expense_report(
            ReportPeriod(str(filters.get("period", ReportPeriod.MONTHLY.value)).upper()),
            start=_as_date(filters.get("start")),
            end=_as_date(filters.get("end")),
        )
        rows: list[list[Any]] = []
        for grouping, group_rows in (
            ("Category", report.by_category),
            ("Vendor", report.by_vendor),
            ("Method", report.by_method),
        ):
            for r in group_rows:
                rows.append(
                    [
                        grouping,
                        r.label,
                        _money(r.total),
                        r.count,
                        float(r.share_of(report.total)),
                    ]
                )
        return Dataset(
            name="expense_summary",
            headers=["Grouped by", "Name", "Total", "Count", "Share %"],
            rows=rows,
        )

    def _profit_loss(self, filters: dict[str, Any]) -> Dataset:
        """The P&L as a label/value table.

        Shaped as rows rather than columns because that is what a P&L reads like on
        paper, and because it then renders correctly through all four exporters
        (CSV/XLSX/JSON/PDF) with no format-specific special-casing.
        """
        report = AccountingService(self.ctx).profit_loss(
            ReportPeriod(str(filters.get("period", ReportPeriod.MONTHLY.value)).upper()),
            start=_as_date(filters.get("start")),
            end=_as_date(filters.get("end")),
        )
        rows: list[list[Any]] = [
            ["Basis", "Cash basis (not an accounting statement)"],
            ["Period start", report.start.isoformat()],
            ["Period end", report.end.isoformat()],
            ["Revenue (collections received)", _money(report.revenue)],
            ["Cost of goods sold", _money(report.cost_of_goods_sold)],
            ["Gross profit", _money(report.gross_profit)],
            ["Operating expenses", _money(report.operating_expenses)],
        ]
        # Indent the category detail under the operating-expenses line so the table
        # reads as a hierarchy in a flat CSV.
        rows.extend(
            [f"    {r.label}", _money(r.total)] for r in report.expenses_by_category
        )
        rows.append(["Net profit", _money(report.net_profit)])
        rows.append(["Net margin %", float(report.net_margin_pct)])
        return Dataset(name="profit_loss", headers=["Item", "Amount"], rows=rows)

    def _cash_flow(self, filters: dict[str, Any]) -> Dataset:
        report = AccountingService(self.ctx).cash_flow(
            ReportPeriod(str(filters.get("period", ReportPeriod.MONTHLY.value)).upper()),
            start=_as_date(filters.get("start")),
            end=_as_date(filters.get("end")),
        )
        rows: list[list[Any]] = [
            [r.label, _money(r.money_in), _money(r.money_out), _money(r.net)]
            for r in report.rows
        ]
        # The totals belong IN the file: a spreadsheet someone forwards should not
        # need the app open to answer "so what was the net".
        rows.append(
            [
                "Total",
                _money(report.total_in),
                _money(report.total_out),
                _money(report.net_flow),
            ]
        )
        return Dataset(
            name="cash_flow",
            headers=["Period", "Money in", "Money out", "Net"],
            rows=rows,
        )

    def _aging_receivable(self, filters: dict[str, Any]) -> Dataset:
        """Point-in-time, so it takes `end` as the as-at date rather than a range."""
        report = AccountingService(self.ctx).aging_receivable(
            as_at=_as_date(filters.get("end"))
        )
        rows = [
            [
                c.name,
                c.phone or "",
                _money(c.buckets.get("CURRENT", Decimal("0"))),
                _money(c.buckets.get("D1_30", Decimal("0"))),
                _money(c.buckets.get("D31_60", Decimal("0"))),
                _money(c.buckets.get("D61_90", Decimal("0"))),
                _money(c.buckets.get("D90_PLUS", Decimal("0"))),
                _money(c.total),
                c.oldest_days,
            ]
            for c in report.customers
        ]
        return Dataset(
            name="aging_receivable",
            headers=[
                "Customer",
                "Phone",
                "Not due yet",
                "1-30 days",
                "31-60 days",
                "61-90 days",
                "90+ days",
                "Total owed",
                "Oldest (days)",
            ],
            rows=rows,
        )

    def _tax_summary(self, filters: dict[str, Any]) -> Dataset:
        report = AccountingService(self.ctx).tax_summary(
            ReportPeriod(str(filters.get("period", ReportPeriod.MONTHLY.value)).upper()),
            start=_as_date(filters.get("start")),
            end=_as_date(filters.get("end")),
        )
        rows: list[list[Any]] = [
            [f"{r.rate}%", _money(r.taxable_base), _money(r.tax_amount), r.line_count]
            for r in report.rows
        ]
        rows.append(["Total", _money(report.total_taxable), _money(report.total_tax), ""])
        if not report.reconciles:
            # Never let the file imply a complete breakdown when it is not one.
            rows.append(
                [
                    "Tax billed on credits (some charged at credit level, not per line)",
                    "",
                    _money(report.total_tax_on_credits),
                    "",
                ]
            )
        return Dataset(
            name="tax_summary",
            headers=["Rate", "Taxable amount", "Tax", "Lines"],
            rows=rows,
        )

    # =================================================================== reads
    def list_exports(self, page: PageInput | None = None) -> Page[ExportJob]:
        self.require(Permission.EXPORT_CREATE)
        stmt = (
            select(ExportJob)
            .where(
                ExportJob.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(ExportJob.deleted_at).is_(None),
            )
            .order_by(col(ExportJob.created_at).desc())
        )
        return paginate(self.session, stmt, page or PageInput())

    def get_export(self, export_id: str) -> ExportJob:
        self.require(Permission.EXPORT_CREATE)
        job = self.session.get(ExportJob, export_id)
        if job is None or job.deleted_at is not None:
            raise NotFoundError("Export not found")
        self.assert_in_scope(job.business_id)  # TENANCY BOUNDARY
        return job

    def download_url(self, export_id: str) -> str | None:
        """A URL for the export's file, or None if it has expired / failed / gone."""
        job = self.get_export(export_id)
        if ExportState(job.state) is not ExportState.READY or not job.file_id:
            return None
        if job.expires_at and utcnow() >= ensure_utc(job.expires_at):
            return None  # the sweep has not run yet, but it is already dead
        return StorageService(self.session).url_for_id(job.file_id)

    # ================================================================== delete
    async def delete_export(self, export_id: str) -> int:
        """Delete one export on the user's say-so: file, row and all.

        The button next to "Download". Someone who has saved the spreadsheet should
        not have to wait 24 hours for their own copy to stop counting against their
        storage quota.

        Permission is EXPORT_CREATE rather than a new ``export:delete``: a caller who
        can generate an export can already replace it at will, so being able to throw
        one away grants no reach they did not have. Adding a permission would mean a
        role migration for no security gain.

        Returns bytes freed. Idempotent from the caller's side only in the sense that
        a second call 404s -- the row is really gone.
        """
        self.require(Permission.EXPORT_CREATE)
        job = self.get_export(export_id)  # TENANCY BOUNDARY (asserts scope)
        return await self._purge(
            self.session,
            job,
            actor_user_id=self.user.id,
            actor_label=self.user.email,
            reason="deleted by user",
        )

    # ============================================================== expiration
    @staticmethod
    async def _purge(
        session: Session,
        job: ExportJob,
        *,
        actor_user_id: str | None = None,
        actor_label: str = "scheduler",
        reason: str,
    ) -> int:
        """Irreversibly remove one export -- bytes, then row -- and audit it.

        The single place an ExportJob dies, so the audit entry can never be skipped
        by a new caller. Returns bytes freed.
        """
        freed = 0
        if job.file_id:
            asset = session.get(FileAsset, job.file_id)
            if asset is not None:
                freed += await StorageService(session).hard_delete(asset)

        session.add(
            AuditLog(
                business_id=job.business_id,
                action=AuditAction.PURGE,
                entity_type="export_job",
                entity_id=job.id,
                summary=(
                    f"Export purged ({reason}): {job.format} of "
                    f"{', '.join(job.datasets) or 'no datasets'}, "
                    f"{job.row_count} row(s), {job.size_bytes} byte(s)"
                ),
                actor_user_id=actor_user_id,
                actor_label=actor_label,
            )
        )

        # Hard delete, not soft: the point of the TTL is that the record stops
        # occupying space. A deleted_at row still costs pages and still leaks what
        # was exported. The AuditLog entry above IS the surviving receipt.
        session.delete(job)
        session.flush()
        return freed

    @staticmethod
    async def purge_stale(
        session: Session, *, business_id: str | None = None, now: datetime | None = None
    ) -> tuple[int, int]:
        """Delete every export past its TTL -- file and row both.

        Returns (jobs_purged, bytes_freed). Called by the hourly purge job with no
        ``business_id`` (all tenants); the Storage Dashboard calls it with one.

        Any state qualifies, not just READY: a FAILED job that nobody will ever
        download is still a row, and a job stuck in RUNNING past its own TTL is a
        crashed run whose temp file needs collecting.
        """
        moment = now or utcnow()
        stmt = select(ExportJob).where(
            col(ExportJob.expires_at).is_not(None),
            col(ExportJob.expires_at) < moment,
        )
        if business_id:
            stmt = stmt.where(ExportJob.business_id == business_id)  # TENANCY BOUNDARY

        purged = freed = 0
        for job in session.execute(stmt).scalars().all():
            freed += await ExportService._purge(session, job, reason="expired")
            purged += 1
        return purged, freed

    # ================================================================= archive
    async def export_archive_batch(self, ctx: ServiceContext, batch_id: str) -> ExportJob:
        """A complete snapshot of everything in an ArchiveBatch.

        Exports ARE the archive-download mechanism (see models/retention.py): the
        owner must be able to take their data with them before the purge, or the
        retention policy is just data loss with extra steps. XLSX, because a batch
        spans several tables and one workbook with a sheet each is the shape a
        non-technical owner can actually open.
        """
        self.require(Permission.RETENTION_MANAGE)

        batch = self.session.get(ArchiveBatch, batch_id)
        if batch is None or batch.deleted_at is not None:
            raise NotFoundError("Archive batch not found")
        self.assert_in_scope(batch.business_id)  # TENANCY BOUNDARY

        business = self.get_business()
        filters: dict[str, Any] = {
            "archive_batch_id": batch_id,
            "include_archived": True,
            "include_voided": True,
        }

        job = ExportJob(
            business_id=self.scope_id,  # TENANCY BOUNDARY
            format=ExportFormat.XLSX,
            state=ExportState.PENDING,
            datasets=["credits", "payments", "customers", "business"],
            filters=filters,
            requested_by_user_id=ctx.user.id if ctx.user else None,
        )
        self.session.add(job)
        self.session.flush()

        try:
            sets = [
                self._credits(filters),
                self._payments(filters),
                self._credit_items(batch_id),
                self._customers({}),
                self._business_dataset({}),
            ]
            bundle = ExportBundle(
                data=self._to_xlsx(sets),
                filename=f"{business.slug}-archive-{batch_id[:8]}-{utcnow():%Y%m%d}.xlsx",
                content_type=_CONTENT_TYPES["xlsx"],
                row_count=sum(s.row_count for s in sets),
            )
            job.state = ExportState.RUNNING
            asset = await StorageService(self.session).upload(
                business_id=self.scope_id,  # TENANCY BOUNDARY
                kind=FileKind.EXPORT,
                filename=bundle.filename,
                data=bundle.data,
                content_type=bundle.content_type,
                user_id=job.requested_by_user_id,
                expires_in_hours=settings.EXPORT_TTL_HOURS,
            )
            StorageService(self.session).attach(asset.id)

            job.file_id = asset.id
            job.row_count = bundle.row_count
            job.size_bytes = len(bundle.data)
            job.state = ExportState.READY
            job.expires_at = utcnow() + timedelta(hours=settings.EXPORT_TTL_HOURS)
            job.completed_at = utcnow()

            batch.export_id = job.id
            self.session.add(batch)

            self.audit(
                AuditAction.EXPORT,
                "archive_batch",
                batch.id,
                f"Exported archive batch {batch.id[:8]} ({bundle.row_count} rows) before purge",
            )
        except Exception as exc:  # noqa: BLE001
            job.state = ExportState.FAILED
            job.error = str(exc)[:1000]
            job.completed_at = utcnow()
            job.file_id = None
            # A FAILED row expires on the same clock as a successful one. Without
            # this it has no expires_at, so purge_stale never matches it and the
            # error sits in the user's exports table forever.
            job.expires_at = utcnow() + timedelta(hours=settings.EXPORT_TTL_HOURS)

        self.session.add(job)
        self.session.flush()
        return job

    def _credit_items(self, batch_id: str) -> Dataset:
        """Line items of the batch's credits -- a snapshot is worthless without them."""
        credit_ids = select(col(Credit.id)).where(
            Credit.business_id == self.scope_id,  # TENANCY BOUNDARY
            Credit.archive_batch_id == batch_id,
        )
        stmt = (
            select(CreditItem, Credit.number)
            .join(Credit, col(CreditItem.credit_id) == col(Credit.id))
            .where(
                CreditItem.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(CreditItem.credit_id).in_(credit_ids),
            )
            .order_by(col(Credit.number).asc(), col(CreditItem.position).asc())
        )
        rows = [
            [
                number,
                i.name,
                i.description,
                str(i.quantity),
                i.unit,
                _money(i.unit_price),
                _money(i.discount_amount),
                _money(i.tax_amount),
                _money(i.line_total),
            ]
            for i, number in self.session.execute(stmt).all()
        ]
        return Dataset(
            name="credit_items",
            headers=[
                "Credit", "Item", "Description", "Qty", "Unit", "Unit price",
                "Discount", "Tax", "Line total",
            ],
            rows=rows,
        )


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _csv_bytes(ds: Dataset) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(ds.headers)
    for row in ds.rows:
        w.writerow([_cell(v) for v in row])
    # utf-8-sig -- Excel needs the BOM or non-ASCII customer names come out garbled.
    return buf.getvalue().encode("utf-8-sig")


def _cell(value: Any) -> Any:
    """Coerce a value into something csv/openpyxl/json can all take."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _money(value: Decimal | None) -> float:
    """Money leaves the system as a NUMBER so the recipient can sum it.

    ``value`` is already a Decimal here -- MoneyType converted the stored integer
    minor units on read, so there is no /100 to do (and doing one would divide by
    100 twice).
    """
    return float(value) if value is not None else 0.0


def _dt(value: datetime | None) -> str:
    return value.isoformat() if value else ""


def _json_safe(value: Any) -> Any:
    """Coerce a filter payload into something a JSON column can actually hold.

    Dates and Decimals are the two types this app passes around constantly and JSON
    knows nothing about. Rendering them as ISO strings / plain strings is lossless
    here, and the readers (_as_date) already accept both.
    """
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _as_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, str):
        return date.fromisoformat(value)
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def _apply_date_filter(stmt: Any, column: Any, filters: dict[str, Any]) -> Any:
    """For DATE columns (issued_date, due_date): already business-local, compare directly."""
    if start := _as_date(filters.get("start")):
        stmt = stmt.where(column >= start)
    if end := _as_date(filters.get("end")):
        stmt = stmt.where(column <= end)
    return stmt


__all__ = ["DATASETS", "Dataset", "ExportBundle", "ExportService"]
