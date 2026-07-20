"""ImportService -- bulk-load customers and credits from a spreadsheet.

WHY THIS EXISTS
---------------
A shop that has been running for years arrives with history: a ledger book typed
into Excel, an export from whatever they used before. Making them retype 400
customers one form at a time is how you lose them on day one. This turns that
sheet into records.

THE THREE RULES THIS MODULE IS BUILT AROUND
--------------------------------------------
R1  VALIDATE EVERYTHING, THEN WRITE -- OR WRITE NOTHING.
    A partial import is the worst outcome there is: the owner cannot tell which
    rows landed, and re-uploading the sheet duplicates the ones that did. So the
    whole file is validated first and a single bad row aborts the batch. Every
    write in ``_commit`` shares one transaction and one commit.

R2  THE IMPORTER IS NOT A SECOND WRITE PATH.
    It calls ``CustomerService.build`` and ``CreditService.create`` -- the same
    code the web form uses. It does not touch a model directly. If it did, every
    invariant those services own (credit totals, customer aggregates, credit
    score, audit trail, stock) would have to be re-implemented here and would rot
    out of sync. This module's job is *parsing and validation*; the services keep
    their job.

R3  ERRORS ARE ADDRESSED TO A SHOPKEEPER, NOT A DEVELOPER.
    Every message names the spreadsheet row, the column heading they can see in
    Excel, and what to do about it. "Row 14, due_date: use YYYY-MM-DD, e.g.
    2026-08-01 (got '14/8/26')" -- not "ValueError: invalid isoformat string".

The dry-run pass is the same code as the real one, minus the commit. That is
deliberate: a preview that runs different logic to the import is a preview that
lies.

TENANCY: every query filters on ``self.scope_id`` (BaseService), and rows are
written through services that do the same. A sheet cannot name another tenant --
there is no business_id column, by design.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlmodel import col, select

from app.core.errors import AppError, ValidationError
from app.core.security import Permission
from app.models.customer import Customer
from app.models.enums import AuditAction, CustomerStatus, ItemKind, PaymentMethod
from app.services.base import BaseService, ServiceContext
from app.services.catalog import CategoryService, ProductService, ServiceItemService
from app.services.credit import CreditItemInput, CreditService
from app.services.customer import CustomerService
from app.services.expense import ExpenseCategoryService, ExpenseService
from app.services.vendor import VendorService
from app.utils.dates import today_in

# Guard rails. A sheet bigger than this is a migration, not an import -- it should
# be chunked, and the owner told so plainly rather than left watching a spinner
# while one request tries to write 50,000 rows inside one transaction.
MAX_IMPORT_ROWS = 2000
MAX_IMPORT_BYTES = 5 * 1024 * 1024

_XLSX_MAGIC = b"PK\x03\x04"  # xlsx is a zip archive

ZERO_MONEY = Decimal("0")


# ---------------------------------------------------------------------------
# Column specification -- the single source of truth
# ---------------------------------------------------------------------------
# The template, the parser, the validator and the field guide the UI renders are
# ALL generated from these tuples. Add a column here and the downloadable template
# grows a heading, the importer starts accepting it, and the help text appears in
# the UI -- with no chance of the four drifting apart.
@dataclass(frozen=True, slots=True)
class Column:
    key: str  # the CSV heading, and the identifier used in error messages
    label: str  # the human heading (accepted as an alias on upload)
    required: bool = False
    help: str = ""
    example: str = ""
    choices: tuple[str, ...] = ()


CUSTOMER_COLUMNS: tuple[Column, ...] = (
    Column(
        "name",
        "Name",
        required=True,
        help="The customer's full name. This is the only column you must fill in.",
        example="Sonam Dorji",
    ),
    Column(
        "phone",
        "Phone",
        help=(
            "Spaces, dashes and + are fine. Worth filling in: it is how a credit "
            "sheet finds this customer later, and where reminders would go."
        ),
        example="+975 17 12 34 56",
    ),
    Column(
        "email",
        "Email",
        help="Where payment reminders are sent. Leave blank if they have no email.",
        example="sonam@example.com",
    ),
    Column("address", "Address", help="Street address.", example="Norzin Lam, Shop 4"),
    Column("city", "City", example="Thimphu"),
    Column(
        "status",
        "Status",
        help="ACTIVE, INACTIVE, BLOCKED or DEFAULTED. Blank means ACTIVE.",
        example="ACTIVE",
        choices=tuple(s.value for s in CustomerStatus),
    ),
    Column(
        "credit_limit",
        "Credit limit",
        help="The most this customer may owe at one time. Blank means no limit.",
        example="5000.00",
    ),
    Column(
        "latitude",
        "Latitude",
        help="Decimal degrees between -90 and 90. Leave blank unless you have a map pin.",
        example="27.4712",
    ),
    Column(
        "longitude",
        "Longitude",
        help="Decimal degrees between -180 and 180.",
        example="89.6339",
    ),
    Column("emergency_contact_name", "Emergency contact name", example="Pema Dorji"),
    Column("emergency_contact_phone", "Emergency contact phone", example="+975 17 65 43 21"),
    Column("emergency_contact_relation", "Emergency contact relation", example="Brother"),
    Column(
        "notes",
        "Notes",
        help="Anything you want to remember about this customer.",
        example="Buys on credit at the end of every month",
    ),
)

CREDIT_COLUMNS: tuple[Column, ...] = (
    Column(
        "customer_code",
        "Customer code",
        help=(
            "The code from your customer list, e.g. CUST-0007. Fill in this column "
            "OR customer_phone. The customer must already exist -- import your "
            "customers first."
        ),
        example="CUST-0007",
    ),
    Column(
        "customer_phone",
        "Customer phone",
        help=(
            "Used only when customer_code is blank. Must match exactly one customer; "
            "if two customers share a phone, use the code instead."
        ),
        example="+975 17 12 34 56",
    ),
    Column(
        "item_name",
        "Item or service",
        required=True,
        help="What was taken on credit. One row is one credit record.",
        example="Rice 5kg",
    ),
    Column(
        "sku",
        "Product SKU",
        help=(
            "Fill this in to link the row to a product you sell -- its stock is "
            "then reduced, exactly as if you had written the credit by hand. Leave "
            "it blank for a one-off item, or when loading OLD credits whose stock "
            "was taken long ago."
        ),
        example="RICE-5",
    ),
    Column("quantity", "Quantity", help="Up to 3 decimals. Blank means 1.", example="2"),
    Column(
        "unit_price",
        "Unit price",
        required=True,
        help="The price of ONE unit, before any discount or tax.",
        example="450.00",
    ),
    Column("unit", "Unit", help="pcs, kg, box, hour... Blank means pcs.", example="kg"),
    Column("item_description", "Item description", example="Local red rice"),
    Column(
        "item_discount_amount",
        "Discount amount",
        help="Money off this credit -- not a percentage. Blank means no discount.",
        example="50.00",
    ),
    Column(
        "issued_date",
        "Issued date",
        help="The day the goods were taken. Format YYYY-MM-DD. Blank means today.",
        example="2026-07-01",
    ),
    Column(
        "due_date",
        "Due date",
        required=True,
        help="When payment is expected. Format YYYY-MM-DD. Cannot be before the issued date.",
        example="2026-08-01",
    ),
    Column(
        "reminder_date",
        "Reminder date",
        help=(
            "When to remind them. Format YYYY-MM-DD, on or before the due date. "
            "Blank lets your reminder settings decide."
        ),
        example="2026-07-25",
    ),
    Column(
        "tax_percentage",
        "Tax %",
        help="0 to 100. Blank uses your business tax rate from Settings.",
        example="0",
    ),
    Column(
        "discount_percentage",
        "Extra discount %",
        help="0 to 100, taken off on top of the discount amount. Blank means none.",
        example="5",
    ),
    Column(
        "initial_payment",
        "Amount already paid",
        help=(
            "How much they have paid against this credit so far. Blank means nothing "
            "paid. Cannot be more than the credit total."
        ),
        example="200.00",
    ),
    Column("notes", "Notes", help="Anything worth remembering.", example="Pays at month end"),
)


PRODUCT_COLUMNS: tuple[Column, ...] = (
    Column(
        "name",
        "Name",
        required=True,
        help="What you call it on the shelf. The only column you must fill in.",
        example="Rice 5kg",
    ),
    Column(
        "price",
        "Selling price",
        help="What you charge the customer for one unit. Blank means 0.",
        example="450.00",
    ),
    Column(
        "unit",
        "Unit",
        help="pcs, kg, litre, box... Blank means pcs.",
        example="kg",
    ),
    Column(
        "category",
        "Category",
        help=(
            "A group like Groceries or Drinks. Type the name -- if it does not exist "
            "yet it will be created for you."
        ),
        example="Groceries",
    ),
    Column(
        "sku",
        "SKU",
        help="Your own code for it. Must be unique in your shop; leave blank if you don't use one.",
        example="RICE-5KG",
    ),
    Column(
        "barcode",
        "Barcode",
        help="The number under the barcode. Must be unique in your shop.",
        example="8901234567890",
    ),
    Column(
        "cost_price",
        "Cost price",
        help="What YOU paid for one unit. Used for profit reports; never shown to a customer.",
        example="380.00",
    ),
    Column(
        "stock_quantity",
        "Stock on hand",
        help="How many you have right now. Up to 3 decimals. Blank means 0.",
        example="24",
    ),
    Column(
        "low_stock_threshold",
        "Low stock alert at",
        help="Warn you when stock drops to this. Blank means no warning.",
        example="5",
    ),
    Column(
        "tax_percentage",
        "Tax %",
        help="0 to 100. Blank uses your business tax rate from Settings.",
        example="0",
    ),
    Column(
        "is_active",
        "Active",
        help="YES or NO. Blank means YES. NO hides it from the sale screen without deleting it.",
        example="YES",
        choices=("YES", "NO"),
    ),
    Column("description", "Description", example="Local red rice, 5kg bag"),
)

SERVICE_COLUMNS: tuple[Column, ...] = (
    Column(
        "name",
        "Name",
        required=True,
        help="What the customer asks for. The only column you must fill in.",
        example="Puncture repair",
    ),
    Column(
        "price",
        "Price",
        help="What you charge for it. Blank means 0.",
        example="150.00",
    ),
    Column(
        "category",
        "Category",
        help="A group like Repairs. Type the name -- it is created if it does not exist.",
        example="Repairs",
    ),
    Column(
        "code",
        "Code",
        help="Your own short code. Must be unique in your shop; blank if you don't use one.",
        example="SVC-PUNC",
    ),
    Column(
        "duration_minutes",
        "How long it takes (minutes)",
        help="Whole minutes. Blank if it varies.",
        example="30",
    ),
    Column(
        "tax_percentage",
        "Tax %",
        help="0 to 100. Blank uses your business tax rate from Settings.",
        example="0",
    ),
    Column(
        "is_active",
        "Active",
        help="YES or NO. Blank means YES. NO hides it without deleting it.",
        example="YES",
        choices=("YES", "NO"),
    ),
    Column("description", "Description", example="Includes tube patch and refit"),
)


VENDOR_COLUMNS: tuple[Column, ...] = (
    Column(
        "name",
        "Name",
        required=True,
        help="Who you pay -- a wholesaler, the landlord, the electricity company.",
        example="Thimphu Wholesale",
    ),
    Column("phone", "Phone", example="+975 17 12 34 56"),
    Column("email", "Email", example="sales@thimphuwholesale.bt"),
    Column("address", "Address", example="Norzin Lam, Thimphu"),
    Column("notes", "Notes", example="Delivers on Tuesdays"),
)

#: Spellings we accept for a payment method beyond the enum's own values.
#: Keyed on the normalised form (upper case, spaces -> underscores).
_METHOD_ALIASES: dict[str, str] = {
    "MOBILE_BANKING": "MOBILE_MONEY",
    "MOBILE_BANK": "MOBILE_MONEY",
    "MBANKING": "MOBILE_MONEY",
}

EXPENSE_COLUMNS: tuple[Column, ...] = (
    Column(
        "expense_date",
        "Date",
        required=True,
        help="The day you paid it. Format YYYY-MM-DD. Cannot be in the future.",
        example="2026-07-01",
    ),
    Column(
        "amount",
        "Amount",
        required=True,
        help="What you paid. Must be more than zero.",
        example="12000.00",
    ),
    Column(
        "category",
        "Category",
        help=(
            "Rent, Utilities, Fuel... Categories are created for you as they appear, "
            "so you do not have to set them up first."
        ),
        example="Rent",
    ),
    Column(
        "vendor_name",
        "Paid to",
        help=(
            "Who got the money. Matched against your suppliers by name; if there is "
            "no such supplier the name is still kept on the expense."
        ),
        example="Thimphu Wholesale",
    ),
    Column(
        "payment_method",
        "Method",
        help=(
            "CASH, BANK_TRANSFER, CARD, MOBILE_BANKING, CHEQUE or OTHER. "
            "Blank means CASH."
        ),
        example="CASH",
    ),
    Column(
        "provider",
        "Bank or wallet",
        help=(
            "Which bank or wallet it went through, when the method alone is not "
            "enough. Anything you type is kept."
        ),
        example="Bank of Bhutan",
    ),
    Column(
        "cash_account",
        "Paid from",
        help=(
            "The name of one of your cash or bank accounts. Must already exist -- "
            "unlike categories, an account carries a balance, so it is never invented."
        ),
        example="Cash drawer",
    ),
    Column("reference", "Reference", help="Bill number, transfer ref.", example="INV-8841"),
    Column("notes", "Notes", example="July rent"),
)


@dataclass(frozen=True, slots=True)
class DatasetSpec:
    name: str
    title: str
    columns: tuple[Column, ...]
    permission: Permission
    intro: str


DATASETS: dict[str, DatasetSpec] = {
    "customers": DatasetSpec(
        name="customers",
        title="Customers",
        columns=CUSTOMER_COLUMNS,
        permission=Permission.CUSTOMER_WRITE,
        intro="One row for each customer. Only Name is required.",
    ),
    "credits": DatasetSpec(
        name="credits",
        title="Credits",
        columns=CREDIT_COLUMNS,
        permission=Permission.CREDIT_WRITE,
        intro=(
            "One row for each credit record. Import your customers first -- every "
            "row has to point at a customer who already exists."
        ),
    ),
    "products": DatasetSpec(
        name="products",
        title="Products",
        columns=PRODUCT_COLUMNS,
        permission=Permission.CATALOG_WRITE,
        intro=(
            "One row for each thing you sell. Only Name is required. Categories are "
            "created for you as they appear."
        ),
    ),
    "services": DatasetSpec(
        name="services",
        title="Services",
        columns=SERVICE_COLUMNS,
        permission=Permission.CATALOG_WRITE,
        intro=(
            "One row for each service you offer. Only Name is required. Categories "
            "are created for you as they appear."
        ),
    ),
    "vendors": DatasetSpec(
        name="vendors",
        title="Suppliers",
        columns=VENDOR_COLUMNS,
        permission=Permission.VENDOR_WRITE,
        intro="One row for each supplier you pay. Only Name is required.",
    ),
    "expenses": DatasetSpec(
        name="expenses",
        title="Expenses",
        columns=EXPENSE_COLUMNS,
        permission=Permission.EXPENSE_WRITE,
        intro=(
            "One row for each thing you paid for. Date and Amount are required. "
            "Categories are created as they appear; suppliers are matched by name."
        ),
    ),
}


def get_spec(dataset: str) -> DatasetSpec:
    spec = DATASETS.get((dataset or "").strip().lower())
    if spec is None:
        raise ValidationError(
            f"Unknown import type '{dataset}'. Allowed: {', '.join(sorted(DATASETS))}",
            field="dataset",
        )
    return spec


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------
@dataclass(slots=True)
class RowIssue:
    """One problem, addressed to a person looking at their spreadsheet.

    ``row`` is the row number as Excel shows it -- header is row 1, so the first
    data row is 2. Off-by-one here means the owner edits the wrong line.
    """

    row: int
    column: str | None
    message: str


@dataclass(slots=True)
class ImportReport:
    dataset: str
    dry_run: bool
    total_rows: int = 0
    created: int = 0
    errors: list[RowIssue] = field(default_factory=list)
    warnings: list[RowIssue] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------
class ImportService(BaseService):
    # -- templates -----------------------------------------------------------
    def template(self, dataset: str, fmt: str) -> tuple[bytes, str, str]:
        """(bytes, filename, content_type) for a blank, fill-in-the-blanks sheet.

        NOTE -- WHY THERE ARE NO EXAMPLE DATA ROWS.
        The obvious design is to ship a sample row and write "delete this row"
        above it. People do not delete it, and then "Sonam Dorji" is a real
        customer in their database. So the sheet carries headings only, and the
        examples live somewhere unimportable: the Instructions sheet in the
        workbook, and the field guide in the UI.
        """
        spec = get_spec(dataset)
        self.require(spec.permission)
        kind = (fmt or "csv").strip().lower()
        if kind == "csv":
            return (
                self._template_csv(spec),
                f"{spec.name}-import-template.csv",
                "text/csv; charset=utf-8",
            )
        if kind in ("xlsx", "excel"):
            return (
                self._template_xlsx(spec),
                f"{spec.name}-import-template.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        raise ValidationError(
            f"Unknown template format '{fmt}'. Allowed: csv, xlsx", field="format"
        )

    @staticmethod
    def _template_csv(spec: DatasetSpec) -> bytes:
        buf = io.StringIO(newline="")
        writer = csv.writer(buf)
        writer.writerow([c.key for c in spec.columns])
        # utf-8-sig: without the BOM, Excel on Windows opens a UTF-8 CSV as
        # cp1252 and mangles every non-ASCII name in the file.
        return buf.getvalue().encode("utf-8-sig")

    @staticmethod
    def _template_xlsx(spec: DatasetSpec) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()

        # Sheet 1 is the data sheet, and it must stay first: the parser picks the
        # first sheet that looks like data, and a stray Instructions sheet in
        # front of it is a confusing failure.
        ws = wb.active
        ws.title = spec.title

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        required_font = Font(color="FDE68A", bold=True)

        for index, column in enumerate(spec.columns, start=1):
            cell = ws.cell(row=1, column=index, value=column.key)
            cell.fill = header_fill
            # Required headings are tinted amber, so "what must I fill in?" is
            # answerable by looking at the sheet rather than reading the docs.
            cell.font = required_font if column.required else header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if column.help:
                cell.comment = _comment(column)
            ws.column_dimensions[get_column_letter(index)].width = max(
                14, min(34, len(column.key) + 6)
            )

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        # A real dropdown for the enum columns. Beats validating "Actve" server-side
        # and sending them back to fix it.
        for index, column in enumerate(spec.columns, start=1):
            if not column.choices:
                continue
            letter = get_column_letter(index)
            rule = DataValidation(
                type="list",
                formula1='"' + ",".join(column.choices) + '"',
                allow_blank=True,
                showDropDown=False,
            )
            rule.error = f"Choose one of: {', '.join(column.choices)}"
            rule.errorTitle = "Not a valid option"
            ws.add_data_validation(rule)
            rule.add(f"{letter}2:{letter}1001")

        # Sheet 2: the manual. Everything the shopkeeper needs, in the file itself
        # -- because the file is what gets emailed to the person doing the typing.
        guide = wb.create_sheet("Instructions")
        guide["A1"] = f"How to fill in the {spec.title} sheet"
        guide["A1"].font = Font(size=14, bold=True)
        guide["A2"] = spec.intro
        guide["A3"] = (
            "Type your data under the headings on the first sheet, starting at row 2. "
            "Leave a column blank if it does not apply. Do not rename or reorder the "
            "headings. Dates must be written as YYYY-MM-DD, e.g. 2026-08-01."
        )
        for cell in ("A2", "A3"):
            guide[cell].alignment = Alignment(wrap_text=True, vertical="top")
        guide.merge_cells("A2:D2")
        guide.merge_cells("A3:D3")
        guide.row_dimensions[3].height = 32

        headings = ("Column", "Required?", "What to put in it", "Example")
        for index, text in enumerate(headings, start=1):
            cell = guide.cell(row=5, column=index, value=text)
            cell.fill = header_fill
            cell.font = header_font

        for offset, column in enumerate(spec.columns):
            row = 6 + offset
            guide.cell(row=row, column=1, value=column.key)
            guide.cell(row=row, column=2, value="Required" if column.required else "Optional")
            guide.cell(row=row, column=3, value=column.help or column.label).alignment = (
                Alignment(wrap_text=True, vertical="top")
            )
            guide.cell(row=row, column=4, value=column.example or "")

        for letter, width in (("A", 26), ("B", 12), ("C", 68), ("D", 22)):
            guide.column_dimensions[letter].width = width
        guide.freeze_panes = "A6"

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # -- import --------------------------------------------------------------
    def run(
        self,
        ctx: ServiceContext,
        *,
        dataset: str,
        filename: str,
        data: bytes,
        dry_run: bool = True,
    ) -> ImportReport:
        """Validate a sheet and, unless ``dry_run``, write it.

        Returns a report either way. The caller commits nothing and rolls back
        nothing: this method owns the transaction, because R1 (all-or-nothing) is
        only meaningful if one function decides.
        """
        spec = get_spec(dataset)
        self.require(spec.permission)

        if not data:
            raise ValidationError("The file is empty", field="file")
        if len(data) > MAX_IMPORT_BYTES:
            raise ValidationError(
                f"That file is {len(data) / 1_048_576:.1f} MB. The limit is "
                f"{MAX_IMPORT_BYTES // 1_048_576} MB -- split it into smaller sheets.",
                field="file",
            )

        rows, unknown = self._parse(spec, filename, data)
        report = ImportReport(dataset=spec.name, dry_run=dry_run, total_rows=len(rows))

        # Not an error: people keep their own working columns in a sheet ("paid?",
        # "remarks for me"). We ignore them -- but we say so, because a silently
        # dropped column is how an owner discovers three weeks later that the
        # importer never read the one they cared about.
        for heading in unknown:
            report.warnings.append(
                RowIssue(
                    1,
                    heading,
                    f"There is no '{heading}' column in this import, so it was ignored. "
                    f"Check the spelling if you expected it to be brought in.",
                )
            )

        if not rows:
            report.errors.append(
                RowIssue(0, None, "There are no data rows under the headings.")
            )
            return report

        if len(rows) > MAX_IMPORT_ROWS:
            report.errors.append(
                RowIssue(
                    0,
                    None,
                    f"That sheet has {len(rows)} rows. The limit is {MAX_IMPORT_ROWS} per "
                    f"import -- split it and upload the parts one after another.",
                )
            )
            return report

        # Pass 1: validate every row into a ready-to-write payload.
        prepared: list[tuple[int, dict[str, Any]]] = []
        validate = _VALIDATORS[spec.name].__get__(self)
        context = self._build_context(spec)
        for line, raw in rows:
            before = len(report.errors)
            payload = validate(line, raw, report, context)
            if len(report.errors) == before and payload is not None:
                prepared.append((line, payload))

        # R1: one bad row stops the batch. Nothing has been written yet -- the
        # validators are pure, they only append to the report.
        if report.errors or dry_run:
            return report

        try:
            report.created = self._commit(ctx, spec, prepared)
        except AppError as exc:
            # A service refused a row that passed validation -- a rule the
            # validator does not know about. Roll the whole batch back (R1) and
            # report it against the row, rather than leaking a 500.
            self.session.rollback()
            report.created = 0
            report.errors.append(
                RowIssue(getattr(exc, "_import_row", 0) or 0, exc.field, exc.message)
            )
            return report

        self.audit(
            AuditAction.CREATE,
            spec.name,
            None,
            f"Bulk imported {report.created} {spec.name} from {filename or 'a spreadsheet'}",
        )
        self.session.commit()
        return report

    def _commit(
        self, ctx: ServiceContext, spec: DatasetSpec, prepared: list[tuple[int, dict[str, Any]]]
    ) -> int:
        """Write every prepared row through the normal services (R2). No commit here.

        Each branch calls the SAME service the web form calls -- ``build`` where the
        service would otherwise commit per row (which would break R1), ``create``
        where it already leaves the transaction to its caller.
        """
        writers: dict[str, Any] = {
            "customers": lambda line, payload: CustomerService(self.ctx).build(**payload),
            "credits": lambda line, payload: CreditService(self.ctx).create(ctx, **payload),
            "products": lambda line, payload: ProductService(self.ctx).build(**payload),
            "services": lambda line, payload: ServiceItemService(self.ctx).build(**payload),
            "vendors": lambda line, payload: VendorService(self.ctx).build(**payload),
            "expenses": lambda line, payload: ExpenseService(self.ctx).build(**payload),
        }
        write = writers[spec.name]

        # Categories are created HERE, not during validation -- a dry run must not
        # write. Done before the rows so every payload has a real category_id.
        if spec.name in ("products", "services"):
            self._apply_categories(prepared)
        elif spec.name == "expenses":
            self._apply_expense_categories(prepared)

        for line, payload in prepared:
            try:
                write(line, payload)
            except AppError as exc:
                # Tag the row so run() can report WHICH line the service refused.
                exc._import_row = line  # type: ignore[attr-defined]
                raise
        return len(prepared)

    # -- parsing -------------------------------------------------------------
    def _parse(
        self, spec: DatasetSpec, filename: str, data: bytes
    ) -> tuple[list[tuple[int, dict[str, str]]], list[str]]:
        """Sheet bytes -> ([(excel_row_number, {column_key: text})], unknown_headings).

        Blank rows are dropped: a sheet that has been edited in Excel is full of
        them, and they are not an error.
        """
        aliases = _alias_map(spec)
        is_xlsx = data[:4] == _XLSX_MAGIC or (filename or "").lower().endswith((".xlsx", ".xlsm"))
        table = self._read_xlsx(data, aliases) if is_xlsx else self._read_csv(data)
        if not table:
            raise ValidationError(
                "That file has no headings in it. Download the template and fill that in.",
                field="file",
            )

        header, *body = table
        mapping: dict[int, str] = {}
        unknown: list[str] = []
        for index, cell in enumerate(header):
            key = aliases.get(_normalise_heading(cell))
            if key is not None:
                mapping[index] = key
            elif cell.strip():
                unknown.append(cell.strip())

        missing = [c.key for c in spec.columns if c.required and c.key not in mapping.values()]
        if missing:
            raise ValidationError(
                f"The sheet is missing required column(s): {', '.join(missing)}. "
                f"Download the template to get the right headings.",
                field="file",
            )

        rows: list[tuple[int, dict[str, str]]] = []
        for offset, cells in enumerate(body):
            # +2: the header is row 1, so the first data row is row 2 in Excel.
            line = offset + 2
            record = {
                key: (cells[index].strip() if index < len(cells) else "")
                for index, key in mapping.items()
            }
            if any(record.values()):
                rows.append((line, record))

        return rows, unknown

    @staticmethod
    def _read_csv(data: bytes) -> list[list[str]]:
        # utf-8-sig strips Excel's BOM if present; latin-1 is the last resort that
        # cannot fail, so a legacy export never dies on one stray byte.
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:  # pragma: no cover - latin-1 decodes anything
            raise ValidationError("Could not read that file as text", field="file")

        sample = text[:4096]
        try:
            # Sniff , ; and tab: a European Excel writes semicolons by default, and
            # the file looks identical to a shopkeeper.
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel  # type: ignore[assignment]
        return [row for row in csv.reader(io.StringIO(text, newline=""), dialect)]

    @staticmethod
    def _read_xlsx(data: bytes, aliases: dict[str, str]) -> list[list[str]]:
        from openpyxl import load_workbook

        try:
            # data_only: read what a formula evaluated to, not "=SUM(A1:A9)".
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises a zoo of exception types
            raise ValidationError(
                "That file is not a readable Excel workbook. Save it as .xlsx or .csv "
                "and try again.",
                field="file",
            ) from exc

        try:
            # WHICH SHEET IS THE DATA? Pick the one whose first row looks most like
            # our headings.
            #
            # NOT "the sheet with the most rows", which is the tempting heuristic and
            # is wrong on our own template: the Instructions sheet is longer than a
            # half-filled data sheet, so a freshly downloaded template would import
            # its own manual. And not "the first sheet" either -- people reorder tabs.
            # Matching on the headings is the thing that actually identifies data.
            best: list[list[str]] = []
            best_score = 0
            for sheet in wb.worksheets:
                table = [
                    [_cell_text(value) for value in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
                table = [row for row in table if any(cell.strip() for cell in row)]
                if not table:
                    continue
                score = sum(1 for cell in table[0] if _normalise_heading(cell) in aliases)
                if score > best_score:
                    best, best_score = table, score
            # Nothing matched anywhere: fall back to the first non-empty sheet so the
            # caller reports "missing required column X" (actionable) rather than
            # "no headings" (which would be a lie -- there are headings, just wrong ones).
            if best_score == 0:
                for sheet in wb.worksheets:
                    table = [
                        [_cell_text(value) for value in row]
                        for row in sheet.iter_rows(values_only=True)
                    ]
                    table = [row for row in table if any(cell.strip() for cell in row)]
                    if table:
                        return table
            return best
        finally:
            wb.close()

    def _build_context(self, spec: DatasetSpec) -> dict[str, Any]:
        """Everything the validators need to look things up, fetched once.

        A per-row query would be N round trips; a shop's customer list fits in
        memory comfortably at the scale this importer is bounded to.
        """
        if spec.name == "customers":
            return {}

        if spec.name == "vendors":
            from app.models.vendor import Vendor

            rows = self.session.exec(
                select(Vendor).where(
                    Vendor.business_id == self.scope_id,  # TENANCY BOUNDARY
                    col(Vendor.deleted_at).is_(None),
                )
            ).all()
            return {
                "existing": {v.name.strip().lower(): v.id for v in rows},
                # Names claimed by earlier ROWS of this same file. The database
                # cannot see these -- they are all still in flight.
                "seen": {},
            }

        if spec.name == "expenses":
            from app.models.cash_account import CashAccount
            from app.models.vendor import Vendor

            vendors = self.session.exec(
                select(Vendor).where(
                    Vendor.business_id == self.scope_id,  # TENANCY BOUNDARY
                    col(Vendor.deleted_at).is_(None),
                )
            ).all()
            accounts = self.session.exec(
                select(CashAccount).where(
                    CashAccount.business_id == self.scope_id,  # TENANCY BOUNDARY
                    col(CashAccount.deleted_at).is_(None),
                )
            ).all()
            return {
                "vendors": {v.name.strip().lower(): v.id for v in vendors},
                "accounts": {a.name.strip().lower(): a.id for a in accounts},
                "today": today_in(self.get_business().timezone),
            }

        if spec.name in ("products", "services"):
            from app.models.catalog import Product, Service

            def taken(model: Any, field: str) -> dict[str, str]:
                """{code: the name using it} -- so the error can say WHICH product
                already owns the SKU, rather than just "duplicate"."""
                rows = self.session.exec(
                    select(model).where(
                        model.business_id == self.scope_id,  # TENANCY BOUNDARY
                        col(model.deleted_at).is_(None),
                    )
                ).all()
                return {
                    getattr(r, field).strip().lower(): r.name
                    for r in rows
                    if getattr(r, field)
                }

            return {
                "codes": {
                    "sku": taken(Product, "sku"),
                    "barcode": taken(Product, "barcode"),
                    "code": taken(Service, "code"),
                },
                # Codes claimed by earlier ROWS of this same file: {code: line}. The
                # database cannot see these -- they are all still in flight.
                "seen": {"sku": {}, "barcode": {}, "code": {}},
            }

        customers = list(
            self.session.exec(
                select(Customer).where(
                    Customer.business_id == self.scope_id,  # TENANCY BOUNDARY
                    col(Customer.deleted_at).is_(None),
                )
            ).all()
        )
        by_phone: dict[str, list[Customer]] = {}
        for customer in customers:
            digits = _phone_key(customer.phone)
            if digits:
                by_phone.setdefault(digits, []).append(customer)
        # Products, for the optional `sku` column that links a row to the catalog
        # and moves its stock. Matched case-insensitively -- nobody types a SKU's
        # casing consistently into a spreadsheet.
        from app.models.catalog import Product

        products = self.session.exec(
            select(Product).where(
                Product.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Product.deleted_at).is_(None),
                col(Product.sku).is_not(None),
            )
        ).all()

        return {
            "by_code": {c.code.upper(): c for c in customers},
            "by_phone": by_phone,
            "products_by_sku": {p.sku.strip().lower(): p for p in products if p.sku},
            "today": today_in(self.get_business().timezone),
        }

    # -- validation: customers -----------------------------------------------
    def _validate_customer(
        self, line: int, raw: dict[str, str], report: ImportReport, _ctx: dict[str, Any]
    ) -> dict[str, Any] | None:
        get = _reader(line, raw, report)

        name = get.text("name", max_length=160, required=True)
        if not name:
            return None

        payload: dict[str, Any] = {"name": name}
        payload["phone"] = get.text("phone", max_length=40)
        payload["email"] = get.email("email")
        payload["address"] = get.text("address", max_length=500)
        payload["city"] = get.text("city", max_length=120)
        payload["notes"] = get.text("notes", max_length=2000)
        payload["status"] = get.choice("status", CustomerStatus) or CustomerStatus.ACTIVE
        payload["credit_limit"] = get.money("credit_limit")
        payload["latitude"] = get.coordinate("latitude", limit=90)
        payload["longitude"] = get.coordinate("longitude", limit=180)
        payload["emergency_contact_name"] = get.text("emergency_contact_name", max_length=160)
        payload["emergency_contact_phone"] = get.text("emergency_contact_phone", max_length=40)
        payload["emergency_contact_relation"] = get.text(
            "emergency_contact_relation", max_length=60
        )

        # A soft duplicate check. NOT an error: families share a phone, and a shop
        # legitimately has two "Karma". But re-uploading yesterday's sheet is a
        # genuinely easy mistake, and this is where it gets caught -- so it is
        # loud, and it does not block.
        if payload["phone"]:
            existing = self.session.exec(
                select(Customer).where(
                    Customer.business_id == self.scope_id,  # TENANCY BOUNDARY
                    Customer.phone == payload["phone"],
                    col(Customer.deleted_at).is_(None),
                )
            ).first()
            if existing is not None:
                report.warnings.append(
                    RowIssue(
                        line,
                        "phone",
                        f"{existing.name} ({existing.code}) already has this phone number. "
                        f"Importing will create a second customer -- remove this row if it "
                        f"is a duplicate.",
                    )
                )
        return payload

    # -- validation: products / services -------------------------------------
    def _validate_product(
        self, line: int, raw: dict[str, str], report: ImportReport, ctx: dict[str, Any]
    ) -> dict[str, Any] | None:
        get = _reader(line, raw, report)

        name = get.text("name", max_length=200, required=True)
        if not name:
            return None

        payload: dict[str, Any] = {"name": name}
        payload["price"] = get.money("price") or ZERO_MONEY
        payload["cost_price"] = get.money("cost_price")
        payload["unit"] = get.text("unit", max_length=20) or "pcs"
        payload["description"] = get.text("description", max_length=2000)
        payload["sku"] = self._unique_code(line, get, ctx, field="sku", label="SKU")
        payload["barcode"] = self._unique_code(line, get, ctx, field="barcode", label="barcode")
        payload["stock_quantity"] = get.quantity("stock_quantity", allow_zero=True) or ZERO_MONEY
        payload["low_stock_threshold"] = get.quantity("low_stock_threshold", allow_zero=True)
        payload["tax_percentage"] = get.percent("tax_percentage")
        payload["is_active"] = get.yes_no("is_active", default=True)
        payload["category_name"] = self._resolve_category(get, ctx)
        return payload

    def _validate_service(
        self, line: int, raw: dict[str, str], report: ImportReport, ctx: dict[str, Any]
    ) -> dict[str, Any] | None:
        get = _reader(line, raw, report)

        name = get.text("name", max_length=200, required=True)
        if not name:
            return None

        payload: dict[str, Any] = {"name": name}
        payload["price"] = get.money("price") or ZERO_MONEY
        payload["description"] = get.text("description", max_length=2000)
        payload["code"] = self._unique_code(line, get, ctx, field="code", label="code")
        payload["duration_minutes"] = get.whole_number("duration_minutes")
        payload["tax_percentage"] = get.percent("tax_percentage")
        payload["is_active"] = get.yes_no("is_active", default=True)
        payload["category_name"] = self._resolve_category(get, ctx)
        return payload

    def _unique_code(
        self, line: int, get: "_Reader", ctx: dict[str, Any], *, field: str, label: str
    ) -> str | None:
        """SKU / barcode / service code: unique per business, INCLUDING within the file.

        The service checks the database. It cannot see that rows 4 and 40 of THIS
        sheet both say RICE-5KG -- they are both still in flight. Catching it here
        means the owner sees both line numbers at once, instead of the batch aborting
        on row 40 with a message about row 4.
        """
        value = get.text(field, max_length=64)
        if not value:
            return None

        existing = ctx["codes"][field].get(value.lower())
        if existing is not None:
            get.report.errors.append(
                RowIssue(
                    line,
                    field,
                    f"{label} '{value}' is already used by '{existing}'. Every {label} "
                    f"has to be unique in your shop -- change it or leave it blank.",
                )
            )
            return None

        seen = ctx["seen"][field]
        if value.lower() in seen:
            get.report.errors.append(
                RowIssue(
                    line,
                    field,
                    f"{label} '{value}' is used twice in this file -- also on row "
                    f"{seen[value.lower()]}. Every {label} has to be unique.",
                )
            )
            return None
        seen[value.lower()] = line
        return value

    @staticmethod
    def _resolve_category(get: "_Reader", ctx: dict[str, Any]) -> str | None:
        """The category NAME, for _commit to resolve. Reads only -- never writes.

        Categories missing from the sheet are created on demand, DELIBERATELY unlike
        the credits import, which refuses a customer it has not seen. The asymmetry
        is the point: inventing a customer invents a person who owes money, so it
        must be deliberate. A category is a label on a shelf -- it carries no money
        and no identity, and making the shopkeeper hand-create ten of them before
        their first import buys nothing.

        But the CREATION happens in _commit, not here. Validators are pure (see the
        module docstring): creating a category during a dry run would write to a
        database the caller was promised nothing would be written to.
        """
        return get.text("category", max_length=120)

    def _apply_categories(self, prepared: list[tuple[int, dict[str, Any]]]) -> None:
        """Turn each row's category NAME into an id, creating the missing ones.

        Runs inside _commit, once the batch is known to be good. Names are matched
        case-insensitively so "Groceries" and "groceries" on two rows do not produce
        two categories.
        """
        categories = self._category_index()
        service = CategoryService(self.ctx)

        for _line, payload in prepared:
            name = payload.pop("category_name", None)
            if not name:
                payload["category_id"] = None
                continue
            key = name.strip().lower()
            if key not in categories:
                categories[key] = service.build(name.strip()).id
            payload["category_id"] = categories[key]

    def _category_index(self) -> dict[str, str]:
        from app.models.catalog import Category

        rows = self.session.exec(
            select(Category).where(
                Category.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(Category.deleted_at).is_(None),
            )
        ).all()
        return {c.name.strip().lower(): c.id for c in rows}

    # -- validation: vendors -------------------------------------------------
    def _validate_vendor(
        self, line: int, raw: dict[str, str], report: ImportReport, ctx: dict[str, Any]
    ) -> dict[str, Any] | None:
        get = _reader(line, raw, report)
        name = get.text("name", max_length=200, required=True)
        if not name:
            return None

        # Duplicates are refused against BOTH the database and earlier rows of this
        # same file -- the second are still in flight and invisible to a query.
        key = name.strip().lower()
        if key in ctx["existing"]:
            report.errors.append(
                RowIssue(line, "name", f"You already have a supplier called '{name}'.")
            )
            return None
        if key in ctx["seen"]:
            report.errors.append(
                RowIssue(
                    line,
                    "name",
                    f"'{name}' is on row {ctx['seen'][key]} of this sheet as well. "
                    f"Each supplier can only appear once.",
                )
            )
            return None
        ctx["seen"][key] = line

        return {
            "name": name,
            "phone": get.text("phone", max_length=40),
            "email": get.email("email"),
            "address": get.text("address", max_length=500),
            "notes": get.text("notes", max_length=1000),
        }

    # -- validation: expenses ------------------------------------------------
    def _validate_expense(
        self, line: int, raw: dict[str, str], report: ImportReport, ctx: dict[str, Any]
    ) -> dict[str, Any] | None:
        get = _reader(line, raw, report)

        when = get.day("expense_date", required=True)
        amount = get.money("amount", required=True)
        if when is None or amount is None:
            return None

        if amount <= Decimal("0"):
            report.errors.append(
                RowIssue(line, "amount", "An expense has to be more than zero.")
            )
            return None
        if when > ctx["today"]:
            report.errors.append(
                RowIssue(
                    line,
                    "expense_date",
                    f"{when} is in the future. An expense records money you have "
                    f"already paid.",
                )
            )
            return None

        method = PaymentMethod.CASH
        raw_method = get.text("payment_method", max_length=20)
        if raw_method:
            key = raw_method.strip().upper().replace(" ", "_")
            # The UI calls it "Mobile banking", so that is what someone types into a
            # spreadsheet. Rejecting the word we ourselves display would be a trap of
            # our own making; the STORED value is still MOBILE_MONEY.
            key = _METHOD_ALIASES.get(key, key)
            try:
                method = PaymentMethod(key)
            except ValueError:
                report.errors.append(
                    RowIssue(
                        line,
                        "payment_method",
                        f"'{raw_method}' is not a payment method. Use one of: "
                        f"{', '.join(m.value for m in PaymentMethod)}.",
                    )
                )
                return None

        # A cash account carries a BALANCE, so it is never invented -- unlike a
        # category, which is only a label. Same asymmetry as customers on the
        # credits import; see _resolve_category.
        account_id: str | None = None
        account_name = get.text("cash_account", max_length=120)
        if account_name:
            account_id = ctx["accounts"].get(account_name.strip().lower())
            if account_id is None:
                report.errors.append(
                    RowIssue(
                        line,
                        "cash_account",
                        f"You have no account called '{account_name}'. Add it under "
                        f"Cash & Bank first, or leave this blank.",
                    )
                )
                return None

        # A supplier IS matched but never invented: an unmatched name is simply kept
        # as text on the expense, which is exactly what vendor_name is for.
        vendor_name = get.text("vendor_name", max_length=200)
        vendor_id = ctx["vendors"].get(vendor_name.strip().lower()) if vendor_name else None

        return {
            "expense_date": when,
            "amount": amount,
            "category_name": self._resolve_category(get, ctx),
            "vendor_id": vendor_id,
            "vendor_name": vendor_name,
            "cash_account_id": account_id,
            "payment_method": method,
            "provider": get.text("provider", max_length=120),
            "reference": get.text("reference", max_length=120),
            "notes": get.text("notes", max_length=1000),
        }

    def _apply_expense_categories(
        self, prepared: list[tuple[int, dict[str, Any]]]
    ) -> None:
        """Same contract as _apply_categories, against the EXPENSE category table.

        Deliberately not shared with the catalog version: they are different tables
        that happen to have the same shape, and merging them would put spending
        buckets in the product-category dropdown. See models/expense.py.
        """
        from app.models.expense import ExpenseCategory

        rows = self.session.exec(
            select(ExpenseCategory).where(
                ExpenseCategory.business_id == self.scope_id,  # TENANCY BOUNDARY
                col(ExpenseCategory.deleted_at).is_(None),
            )
        ).all()
        index = {c.name.strip().lower(): c.id for c in rows}
        service = ExpenseCategoryService(self.ctx)

        for _line, payload in prepared:
            name = payload.pop("category_name", None)
            if not name:
                payload["category_id"] = None
                continue
            key = name.strip().lower()
            if key not in index:
                index[key] = service.build(name.strip()).id
            payload["category_id"] = index[key]

    # -- validation: credits -------------------------------------------------
    def _validate_credit(
        self, line: int, raw: dict[str, str], report: ImportReport, ctx: dict[str, Any]
    ) -> dict[str, Any] | None:
        get = _reader(line, raw, report)

        customer = self._resolve_customer(line, raw, report, ctx)
        item_name = get.text("item_name", max_length=200, required=True)
        unit_price = get.money("unit_price", required=True)
        quantity = get.quantity("quantity") or Decimal("1")

        if customer is None or not item_name or unit_price is None:
            return None

        issued = get.day("issued_date") or ctx["today"]
        due = get.day("due_date", required=True)
        if due is None:
            return None
        if due < issued:
            report.errors.append(
                RowIssue(
                    line,
                    "due_date",
                    f"The due date ({due}) is before the issued date ({issued}). "
                    f"Check both columns.",
                )
            )
            return None

        reminder = get.day("reminder_date")
        if reminder is not None and reminder > due:
            report.errors.append(
                RowIssue(
                    line,
                    "reminder_date",
                    f"The reminder date ({reminder}) is after the due date ({due}). "
                    f"A reminder has to come first.",
                )
            )
            return None

        discount = get.money("item_discount_amount") or Decimal("0")
        item_total = (unit_price * quantity).quantize(Decimal("0.01"))
        if discount > item_total:
            report.errors.append(
                RowIssue(
                    line,
                    "item_discount_amount",
                    f"The discount ({discount}) is more than the credit is worth "
                    f"({item_total}). This column is an amount of money, not a percentage.",
                )
            )
            return None

        # LINKING TO THE CATALOG IS OPT-IN, PER ROW.
        #
        # A blank `sku` keeps the original behaviour: a CUSTOM line, tied to no
        # product, moving no stock. That is what you want when backfilling a paper
        # ledger -- re-deducting stock for sales that happened months ago would
        # leave every count wrong by the size of your own history.
        #
        # Filling `sku` in says "this is a sale I am registering now": the row is
        # linked to the product and CreditService.create decrements it, exactly as
        # the credit form does.
        product = None
        sku = get.text("sku", max_length=64)
        if sku:
            product = ctx["products_by_sku"].get(sku.strip().lower())
            if product is None:
                report.errors.append(
                    RowIssue(
                        line,
                        "sku",
                        f"You have no product with the SKU '{sku}'. Check the "
                        f"spelling, or leave the column blank to record this as a "
                        f"one-off item that does not touch stock.",
                    )
                )
                return None

        item = CreditItemInput(
            name=item_name,
            quantity=quantity,
            unit_price=unit_price,
            kind=ItemKind.PRODUCT if product else ItemKind.CUSTOM,
            product_id=product.id if product else None,
            description=get.text("item_description", max_length=500),
            unit=get.text("unit", max_length=20) or "pcs",
            discount_amount=discount,
            tax_percentage=Decimal("0"),  # tax is credit-level for a one-item sheet
        )

        payload: dict[str, Any] = {
            "customer_id": customer.id,
            "items": [item],
            "issued_date": issued,
            "due_date": due,
            "reminder_date": reminder,
            "tax_percentage": get.percent("tax_percentage"),
            "discount_percentage": get.percent("discount_percentage"),
            "notes": get.text("notes", max_length=2000),
            "initial_payment": get.money("initial_payment"),
        }

        paid = payload["initial_payment"]
        if paid is not None:
            # PaymentService would refuse an overpayment mid-batch and abort the
            # whole import (R1). Catching it here means the owner sees every bad
            # row at once instead of fixing them one upload at a time.
            after_discounts = item_total - discount
            pct = payload["discount_percentage"]
            if pct:
                after_discounts -= (after_discounts * pct / Decimal("100")).quantize(
                    Decimal("0.01")
                )
            tax_pct = payload["tax_percentage"] or Decimal("0")
            estimated = after_discounts + (after_discounts * tax_pct / Decimal("100")).quantize(
                Decimal("0.01")
            )
            if paid > estimated:
                report.errors.append(
                    RowIssue(
                        line,
                        "initial_payment",
                        f"They cannot have paid {paid} against a credit worth about "
                        f"{estimated.quantize(Decimal('0.01'))}. Lower the amount, or "
                        f"check the price and quantity.",
                    )
                )
                return None
        return payload

    def _resolve_customer(
        self, line: int, raw: dict[str, str], report: ImportReport, ctx: dict[str, Any]
    ) -> Customer | None:
        code = (raw.get("customer_code") or "").strip().upper()
        phone = (raw.get("customer_phone") or "").strip()

        if code:
            customer = ctx["by_code"].get(code)
            if customer is None:
                report.errors.append(
                    RowIssue(
                        line,
                        "customer_code",
                        f"No customer has the code {code}. Check your customer list, or "
                        f"import the customer first.",
                    )
                )
                return None
        elif phone:
            matches = ctx["by_phone"].get(_phone_key(phone), [])
            if not matches:
                report.errors.append(
                    RowIssue(
                        line,
                        "customer_phone",
                        f"No customer has the phone number {phone}. Import the customer "
                        f"first, then import this sheet.",
                    )
                )
                return None
            if len(matches) > 1:
                names = ", ".join(f"{c.name} ({c.code})" for c in matches[:3])
                report.errors.append(
                    RowIssue(
                        line,
                        "customer_phone",
                        f"{len(matches)} customers share the phone {phone}: {names}. "
                        f"Use the customer_code column instead so there is no doubt.",
                    )
                )
                return None
            customer = matches[0]
        else:
            report.errors.append(
                RowIssue(
                    line,
                    "customer_code",
                    "Fill in either customer_code or customer_phone so we know whose "
                    "credit this is.",
                )
            )
            return None

        if CustomerStatus(customer.status) is CustomerStatus.BLOCKED:
            report.errors.append(
                RowIssue(
                    line,
                    "customer_code",
                    f"{customer.name} ({customer.code}) is blocked from taking credit. "
                    f"Set them to Active first, or remove this row.",
                )
            )
            return None
        return customer


# ---------------------------------------------------------------------------
# Cell readers -- each one records a friendly error and returns None on bad input
# ---------------------------------------------------------------------------
#: dataset -> the validator that turns one row into a ready-to-write payload.
#: A dict rather than an if/elif chain: adding a dataset is a spec plus a function,
#: and get_spec already rejects any name that is not in DATASETS.
_VALIDATORS = {
    "customers": ImportService._validate_customer,
    "credits": ImportService._validate_credit,
    "products": ImportService._validate_product,
    "services": ImportService._validate_service,
    "vendors": ImportService._validate_vendor,
    "expenses": ImportService._validate_expense,
}


def _reader(line: int, raw: dict[str, str], report: ImportReport) -> "_Reader":
    return _Reader(line, raw, report)


class _Reader:
    """Reads one row's cells, appending an issue instead of raising.

    Raising on the first bad cell would show the owner one error per upload. The
    whole point of the preview is to show them all of them at once, so every
    reader here degrades to None and moves on.
    """

    __slots__ = ("line", "raw", "report")

    def __init__(self, line: int, raw: dict[str, str], report: ImportReport) -> None:
        self.line = line
        self.raw = raw
        self.report = report

    def _fail(self, column: str, message: str) -> None:
        self.report.errors.append(RowIssue(self.line, column, message))

    def _cell(self, column: str) -> str:
        return (self.raw.get(column) or "").strip()

    def text(self, column: str, *, max_length: int, required: bool = False) -> str | None:
        value = self._cell(column)
        if not value:
            if required:
                self._fail(column, f"{column} is required and this row leaves it blank.")
            return None
        if len(value) > max_length:
            self._fail(
                column,
                f"{column} is {len(value)} characters; the most it can be is {max_length}.",
            )
            return None
        return value

    def email(self, column: str) -> str | None:
        value = self._cell(column)
        if not value:
            return None
        if not _EMAIL_RE.match(value):
            self._fail(column, f"'{value}' is not a valid email address.")
            return None
        return value.lower()

    def choice(self, column: str, enum: type[CustomerStatus]) -> CustomerStatus | None:
        value = self._cell(column)
        if not value:
            return None
        try:
            return enum(value.strip().upper())
        except ValueError:
            allowed = ", ".join(m.value for m in enum)
            self._fail(column, f"'{value}' is not a valid {column}. Use one of: {allowed}.")
            return None

    def money(self, column: str, *, required: bool = False) -> Decimal | None:
        value = _strip_money(self._cell(column))
        if not value:
            if required:
                self._fail(column, f"{column} is required and this row leaves it blank.")
            return None
        if not _MONEY_RE.match(value):
            self._fail(
                column,
                f"'{self._cell(column)}' is not an amount of money. Write it like 450 or "
                f"450.50, with no currency symbol.",
            )
            return None
        amount = Decimal(value)
        if amount < 0:
            self._fail(column, f"{column} cannot be negative.")
            return None
        return amount

    def quantity(self, column: str, *, allow_zero: bool = False) -> Decimal | None:
        """A count. ``allow_zero`` for stock levels -- "I have none left" is a fact,
        whereas selling someone zero of something is a typo."""
        value = _strip_money(self._cell(column))
        if not value:
            return None
        if not _QUANTITY_RE.match(value):
            self._fail(
                column,
                f"'{self._cell(column)}' is not a quantity. Write it like 2 or 1.5.",
            )
            return None
        amount = Decimal(value)
        if amount < 0 or (amount == 0 and not allow_zero):
            self._fail(
                column,
                "Quantity has to be zero or more."
                if allow_zero
                else "Quantity has to be more than zero.",
            )
            return None
        return amount

    def whole_number(self, column: str) -> int | None:
        value = _strip_money(self._cell(column))
        if not value:
            return None
        try:
            number = int(Decimal(value))
        except (InvalidOperation, ValueError):
            self._fail(column, f"'{self._cell(column)}' is not a whole number, like 30.")
            return None
        if number < 0:
            self._fail(column, f"{column} cannot be negative.")
            return None
        return number

    def yes_no(self, column: str, *, default: bool) -> bool:
        """YES/NO, and every spelling of it a real sheet contains.

        Excel turns a checkbox into TRUE, a human types y, and an export writes 1.
        All three mean the same thing, and refusing them would be pedantry.
        """
        value = self._cell(column).strip().lower()
        if not value:
            return default
        if value in ("yes", "y", "true", "1", "active"):
            return True
        if value in ("no", "n", "false", "0", "inactive"):
            return False
        self._fail(column, f"'{self._cell(column)}' is not clear. Write YES or NO.")
        return default

    def percent(self, column: str) -> Decimal | None:
        value = _strip_money(self._cell(column)).rstrip("%")
        if not value:
            return None
        try:
            amount = Decimal(value)
        except InvalidOperation:
            self._fail(column, f"'{self._cell(column)}' is not a percentage. Write it like 5 or 7.5.")
            return None
        if not (0 <= amount <= 100):
            self._fail(column, f"{column} has to be between 0 and 100 (got {amount}).")
            return None
        return amount

    def coordinate(self, column: str, *, limit: int) -> float | None:
        value = self._cell(column)
        if not value:
            return None
        try:
            number = float(value)
        except ValueError:
            self._fail(column, f"'{value}' is not a number. {column} looks like 27.4712.")
            return None
        if not (-limit <= number <= limit):
            self._fail(column, f"{column} has to be between -{limit} and {limit}.")
            return None
        return number

    def day(self, column: str, *, required: bool = False) -> date | None:
        value = self._cell(column)
        if not value:
            if required:
                self._fail(
                    column, f"{column} is required. Write it as YYYY-MM-DD, e.g. 2026-08-01."
                )
            return None
        parsed = _parse_date(value)
        if parsed is None:
            # Deliberately NOT guessing at 08/07/2026. That is the 8th of July to
            # half the world and the 7th of August to the other half, and guessing
            # wrong silently sets a due date a month out.
            self._fail(
                column,
                f"'{value}' is not a date we can read. Write it as YYYY-MM-DD, "
                f"e.g. 2026-08-01.",
            )
            return None
        return parsed


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_MONEY_RE = re.compile(r"^\d{1,12}(\.\d{1,2})?$")
_QUANTITY_RE = re.compile(r"^\d{1,9}(\.\d{1,3})?$")
_NON_DIGIT = re.compile(r"\D")
# Finds the first number inside a formatted cell: "Nu. 1,200.50" -> "1,200.50".
_NUMBER_IN_TEXT = re.compile(r"-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d+(?:\.\d+)?")


def _strip_money(value: str) -> str:
    """'Nu. 1,200.00' -> '1200.00'.  'abc' -> 'abc'.

    Real sheets carry thousands separators and currency symbols, because a human
    formatted the column to look right. Refusing those would fail rows that are
    perfectly unambiguous.

    Text with no number in it comes back UNCHANGED rather than emptied, and that
    is the important half. The readers treat "" as "this cell is blank" and skip
    it, so scrubbing 'abc' down to '' would file a garbage cell as a deliberate
    omission and import the row anyway -- silently wrong, which is the one outcome
    this module exists to prevent.
    """
    if not value:
        return ""
    # \xa0 is a non-breaking space: Excel and pasted-from-a-website cells are full
    # of them, and they are invisible to the person wondering why the row failed.
    compact = value.replace(" ", "").replace("\xa0", "")
    match = _NUMBER_IN_TEXT.search(compact)
    if match is None:
        return value.strip()
    return match.group(0).replace(",", "")


def _phone_key(phone: str | None) -> str:
    """Digits only, so '+975 17-12-34-56' and '9751712 3456' are the same number."""
    return _NON_DIGIT.sub("", phone or "")


def _parse_date(value: str) -> date | None:
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(value.strip(), pattern).date()
        except ValueError:
            continue
    return None


def _cell_text(value: Any) -> str:
    """An Excel cell -> the text a CSV would have carried.

    Excel hands back real types: a date cell is a datetime, and a quantity of 2 is
    the float 2.0. Rendering that with str() gives '2026-08-01 00:00:00' and '2.0',
    which then fail our own validators.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # 2.0 -> '2', but 1.5 -> '1.5'. normalize() alone would give '2E+0'.
        dec = Decimal(str(value)).normalize()
        return format(dec, "f")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def _normalise_heading(value: str) -> str:
    """'Emergency contact name ' / 'Emergency-Contact-Name' -> 'emergency_contact_name'."""
    return re.sub(r"[\s\-]+", "_", (value or "").strip().lower()).strip("_")


def _alias_map(spec: DatasetSpec) -> dict[str, str]:
    """Accept both the machine heading and the human label for every column.

    Someone will retype the headings by hand, or translate them halfway. Matching
    'Credit limit' as well as 'credit_limit' costs one dict and saves a support
    conversation.
    """
    aliases: dict[str, str] = {}
    for column in spec.columns:
        aliases[_normalise_heading(column.key)] = column.key
        aliases[_normalise_heading(column.label)] = column.key
    return aliases


def _comment(column: Column):
    from openpyxl.comments import Comment

    body = column.help or column.label
    if column.example:
        body = f"{body}\n\nExample: {column.example}"
    if column.required:
        body = f"REQUIRED\n\n{body}"
    comment = Comment(body, "Credit System")
    comment.width = 320
    comment.height = 140
    return comment
