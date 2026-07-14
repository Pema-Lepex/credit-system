"""ReportService -- period reports, and the two documents a shop hands a customer.

TWO DIFFERENT THINGS LIVE HERE, ON PURPOSE
------------------------------------------
* ``generate()`` + ``to_csv/to_xlsx/to_pdf`` -- an analytical report over a period.
* ``invoice_pdf()`` / ``receipt_pdf()`` -- a single, printable A4 document.

They share the money helpers, the business's currency, and the timezone discipline,
so splitting them across two modules would mean duplicating all three.

WHY NOTHING HERE IS EVER WRITTEN TO STORAGE
-------------------------------------------
Spec: "do not permanently store generated PDFs", "generate invoices and receipts
only when the user requests a download".

Every renderer below builds its bytes in a ``BytesIO`` and RETURNS them. Nothing in
this module calls ``StorageService.upload``, creates a ``FileAsset``, or touches the
filesystem. A generated document is a *view* of the data, not new data: persisting
it would (a) fill a free-tier disk with regenerable artefacts and (b) let a stale
PDF outlive a correction made to the underlying credit -- a receipt that disagrees
with the ledger is worse than no receipt. The only thing StorageService is used for
here is READING the business logo back out.

MONEY: all sums come from app.services.analytics.money_sum / to_money, which handle
the integer-minor-unit storage of MoneyType. Do not hand-roll a SUM in this file.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import and_, func, select
from sqlmodel import col

from app.core.errors import NotFoundError, ValidationError
from app.core.security import Permission
from app.models.base import utcnow
from app.models.business import Business
from app.models.credit import Credit, CreditItem, Payment
from app.models.customer import Customer
from app.models.enums import CreditStatus, FileKind, PaymentMethod, ReportPeriod
from app.models.file import FileAsset
from app.models.types import quantize_money
from app.services.analytics import count_if, money_sum, money_sum_if, to_money
from app.services.base import BaseService
from app.storage.service import StorageService
from app.utils.dates import ensure_utc, get_tz, period_bounds, start_of_day, today_in

ZERO = Decimal("0")

Granularity = Literal["day", "month"]

# A CUSTOM range longer than this is rendered as months, not days. 62 days of rows is
# already a scroll; 400 would be unreadable and would make the XLSX pointlessly large.
_DAILY_ROW_LIMIT = 62


# ---------------------------------------------------------------------------
# Report shapes
# ---------------------------------------------------------------------------
@dataclass(frozen=True, slots=True)
class ReportRow:
    """One bucket (a day or a month) of the period breakdown."""

    bucket: date
    label: str
    credits_issued: Decimal
    credit_count: int
    collected: Decimal
    payment_count: int

    @property
    def net(self) -> Decimal:
        """Issued minus collected: how much the shop's receivables moved that day."""
        return quantize_money(self.credits_issued - self.collected)


@dataclass(frozen=True, slots=True)
class ReportSummary:
    credits_issued: Decimal
    credit_count: int
    collected: Decimal
    payment_count: int
    outstanding_at_end: Decimal   # receivables as they stood at the period's close
    overdue_amount: Decimal
    overdue_count: int

    @property
    def collection_rate_pct(self) -> Decimal:
        if self.credits_issued == ZERO:
            return ZERO
        return quantize_money(self.collected / self.credits_issued * Decimal("100"))


@dataclass(frozen=True, slots=True)
class MethodBreakdown:
    method: PaymentMethod
    total: Decimal
    count: int


@dataclass(frozen=True, slots=True)
class StatusBreakdown:
    status: CreditStatus
    count: int
    total: Decimal
    outstanding: Decimal


@dataclass(frozen=True, slots=True)
class ReportCustomer:
    customer_id: str
    code: str
    name: str
    credits_issued: Decimal
    credit_count: int
    collected: Decimal
    outstanding: Decimal


@dataclass(frozen=True, slots=True)
class ReportData:
    business_name: str
    currency: str
    currency_symbol: str
    timezone: str

    period: ReportPeriod
    start: date              # inclusive, business-local
    end: date                # inclusive, business-local
    granularity: Granularity
    generated_at: datetime   # UTC

    summary: ReportSummary
    rows: list[ReportRow] = field(default_factory=list)
    top_customers: list[ReportCustomer] = field(default_factory=list)
    by_method: list[MethodBreakdown] = field(default_factory=list)
    by_status: list[StatusBreakdown] = field(default_factory=list)

    @property
    def title(self) -> str:
        return f"{self.period.value.title()} report - {self.start:%d %b %Y} to {self.end:%d %b %Y}"


class ReportService(BaseService):
    # ================================================================ generate
    def generate(
        self,
        period: ReportPeriod,
        start: date | None = None,
        end: date | None = None,
    ) -> ReportData:
        """Build a period report. All bounds are resolved in the business's timezone."""
        self.require(Permission.REPORT_READ)
        business = self.get_business()
        tz = business.timezone

        if period is ReportPeriod.CUSTOM and (start is None or end is None):
            raise ValidationError(
                "A custom report needs both a start and an end date", field="period"
            )
        if start and end and end < start:
            raise ValidationError("The end date is before the start date", field="end")

        # period_bounds gives UTC instants for a half-open [start, end) range anchored
        # on local calendar days -- the single source of truth for "what is in scope".
        lower, upper = period_bounds(period, tz, start=start, end=end)
        local_start = lower.astimezone(get_tz(tz)).date()
        # upper is exclusive local midnight, so the last INCLUDED local day is upper-1.
        local_end = (upper.astimezone(get_tz(tz)) - timedelta(microseconds=1)).date()

        buckets, granularity = self._buckets(local_start, local_end, period)

        return ReportData(
            business_name=business.name,
            currency=business.currency,
            currency_symbol=business.currency_symbol,
            timezone=tz,
            period=period,
            start=local_start,
            end=local_end,
            granularity=granularity,
            generated_at=utcnow(),
            summary=self._summary(local_start, local_end, lower, upper, tz),
            rows=self._rows(buckets, granularity, tz),
            top_customers=self._top_customers(local_start, local_end, lower, upper),
            by_method=self._by_method(lower, upper),
            by_status=self._by_status(local_start, local_end),
        )

    @staticmethod
    def _buckets(
        start: date, end: date, period: ReportPeriod
    ) -> tuple[list[date], Granularity]:
        span_days = (end - start).days + 1
        by_month = period is ReportPeriod.YEARLY or span_days > _DAILY_ROW_LIMIT

        if by_month:
            out: list[date] = []
            cursor = start.replace(day=1)
            while cursor <= end:
                out.append(cursor)
                cursor = (
                    cursor.replace(year=cursor.year + 1, month=1)
                    if cursor.month == 12
                    else cursor.replace(month=cursor.month + 1)
                )
            return out, "month"

        # Every day in range, INCLUDING days with no activity -- same reasoning as
        # month_range in utils/dates: a gap in a report reads as "we don't know".
        return [start + timedelta(days=i) for i in range(span_days)], "day"

    def _summary(
        self, local_start: date, local_end: date, lower: datetime, upper: datetime, tz: str
    ) -> ReportSummary:
        # Issued: keyed on the local calendar date the credit was written.
        credit_stmt = select(
            money_sum(Credit.grand_total).label("issued"),
            func.count().label("count"),
        ).where(
            Credit.business_id == self.scope_id,  # tenancy boundary
            col(Credit.deleted_at).is_(None),
            col(Credit.status) != CreditStatus.CANCELLED,
            col(Credit.issued_date) >= local_start,
            col(Credit.issued_date) <= local_end,
        )
        c_row = self.session.execute(credit_stmt).one()

        # Collected: keyed on the payment INSTANT, so the UTC bounds are what matter.
        pay_stmt = select(
            money_sum(Payment.amount).label("collected"),
            func.count().label("count"),
        ).where(
            Payment.business_id == self.scope_id,  # tenancy boundary
            col(Payment.deleted_at).is_(None),
            col(Payment.voided_at).is_(None),
            col(Payment.paid_at) >= lower,
            col(Payment.paid_at) < upper,
        )
        p_row = self.session.execute(pay_stmt).one()

        # Outstanding AT THE PERIOD END: everything billed on or before the last day,
        # minus everything collected before the period's closing instant. Reconstructed
        # rather than read off remaining_amount, which only knows about *now*.
        billed_stmt = select(
            money_sum_if(
                Credit.grand_total,
                and_(
                    col(Credit.status) != CreditStatus.CANCELLED,
                    col(Credit.issued_date) <= local_end,
                ),
            )
        ).where(
            Credit.business_id == self.scope_id,  # tenancy boundary
            col(Credit.deleted_at).is_(None),
        )
        billed = int(self.session.execute(billed_stmt).scalar_one() or 0)

        collected_ever_stmt = select(money_sum(Payment.amount)).where(
            Payment.business_id == self.scope_id,  # tenancy boundary
            col(Payment.deleted_at).is_(None),
            col(Payment.voided_at).is_(None),
            col(Payment.paid_at) < upper,
        )
        collected_ever = int(self.session.execute(collected_ever_stmt).scalar_one() or 0)

        # Overdue as at the period end -- but never later than TODAY. Running the
        # current month's report on the 14th must not declare a credit due on the 28th
        # overdue just because the 28th falls inside the reporting window: it has not
        # happened yet, and the customer may well pay on time. Clamping the reference
        # date is what makes an in-progress period report truthful.
        # ``<`` not ``<=``: a credit due today is due, not late (matches
        # CreditService._derive_status).
        as_at = min(local_end, today_in(tz))
        late = col(Credit.due_date) < as_at
        overdue_stmt = select(
            count_if(late).label("count"),
            money_sum_if(Credit.remaining_amount, late).label("amount"),
        ).where(
            Credit.business_id == self.scope_id,  # tenancy boundary
            col(Credit.deleted_at).is_(None),
            col(Credit.status).in_(list(CreditStatus.open_statuses())),
            Credit.remaining_amount > 0,
        )
        o_row = self.session.execute(overdue_stmt).one()

        return ReportSummary(
            credits_issued=to_money(c_row.issued),
            credit_count=int(c_row.count or 0),
            collected=to_money(p_row.collected),
            payment_count=int(p_row.count or 0),
            outstanding_at_end=max(ZERO, to_money(billed) - to_money(collected_ever)),
            overdue_amount=to_money(o_row.amount),
            overdue_count=int(o_row.count or 0),
        )

    def _rows(
        self, buckets: list[date], granularity: Granularity, tz: str
    ) -> list[ReportRow]:
        """One SUM(CASE) column set per bucket -- two queries total, not 2*N."""
        if not buckets:
            return []

        edges = [(b, _bucket_end(b, granularity)) for b in buckets]

        credit_cols: list[Any] = []
        for lo, hi in edges:
            window = and_(col(Credit.issued_date) >= lo, col(Credit.issued_date) < hi)
            credit_cols.append(money_sum_if(Credit.grand_total, window))
            credit_cols.append(count_if(window))
        credit_row = self.session.execute(
            select(*credit_cols).where(
                Credit.business_id == self.scope_id,  # tenancy boundary
                col(Credit.deleted_at).is_(None),
                col(Credit.status) != CreditStatus.CANCELLED,
            )
        ).one()

        pay_cols: list[Any] = []
        for lo, hi in edges:
            # Instants: local bucket edges expressed as UTC.
            window = and_(
                col(Payment.paid_at) >= start_of_day(lo, tz),
                col(Payment.paid_at) < start_of_day(hi, tz),
            )
            pay_cols.append(money_sum_if(Payment.amount, window))
            pay_cols.append(count_if(window))
        pay_row = self.session.execute(
            select(*pay_cols).where(
                Payment.business_id == self.scope_id,  # tenancy boundary
                col(Payment.deleted_at).is_(None),
                col(Payment.voided_at).is_(None),
            )
        ).one()

        fmt = "%Y-%m-%d" if granularity == "day" else "%Y-%m"
        return [
            ReportRow(
                bucket=bucket,
                label=bucket.strftime(fmt),
                credits_issued=to_money(credit_row[i * 2]),
                credit_count=int(credit_row[i * 2 + 1] or 0),
                collected=to_money(pay_row[i * 2]),
                payment_count=int(pay_row[i * 2 + 1] or 0),
            )
            for i, bucket in enumerate(buckets)
        ]

    def _top_customers(
        self, local_start: date, local_end: date, lower: datetime, upper: datetime, limit: int = 10
    ) -> list[ReportCustomer]:
        """Top customers BY ACTIVITY IN THE PERIOD.

        Not the cached Customer.total_credit (that is all-time) -- a period report
        that ranked customers by lifetime volume would answer a question nobody asked.
        Two GROUP BY queries, merged on customer_id.
        """
        issued_stmt = (
            select(
                col(Credit.customer_id),
                money_sum(Credit.grand_total).label("issued"),
                func.count().label("count"),
            )
            .where(
                Credit.business_id == self.scope_id,  # tenancy boundary
                col(Credit.deleted_at).is_(None),
                col(Credit.status) != CreditStatus.CANCELLED,
                col(Credit.issued_date) >= local_start,
                col(Credit.issued_date) <= local_end,
            )
            .group_by(col(Credit.customer_id))
        )
        paid_stmt = (
            select(col(Payment.customer_id), money_sum(Payment.amount).label("paid"))
            .where(
                Payment.business_id == self.scope_id,  # tenancy boundary
                col(Payment.deleted_at).is_(None),
                col(Payment.voided_at).is_(None),
                col(Payment.paid_at) >= lower,
                col(Payment.paid_at) < upper,
            )
            .group_by(col(Payment.customer_id))
        )

        issued = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in self.session.execute(issued_stmt)}
        paid = {r[0]: int(r[1] or 0) for r in self.session.execute(paid_stmt)}

        ids = set(issued) | set(paid)
        if not ids:
            return []

        customers = {
            c.id: c
            for c in self.session.execute(
                select(Customer).where(
                    Customer.business_id == self.scope_id,  # tenancy boundary
                    col(Customer.id).in_(ids),
                )
            )
            .scalars()
            .all()
        }

        out = [
            ReportCustomer(
                customer_id=cid,
                code=customers[cid].code,
                name=customers[cid].name,
                credits_issued=to_money(issued.get(cid, (0, 0))[0]),
                credit_count=issued.get(cid, (0, 0))[1],
                collected=to_money(paid.get(cid, 0)),
                outstanding=quantize_money(customers[cid].outstanding_balance),
            )
            for cid in ids
            if cid in customers
        ]
        out.sort(key=lambda c: c.credits_issued, reverse=True)
        return out[:limit]

    def _by_method(self, lower: datetime, upper: datetime) -> list[MethodBreakdown]:
        stmt = (
            select(
                col(Payment.method),
                money_sum(Payment.amount).label("total"),
                func.count().label("count"),
            )
            .where(
                Payment.business_id == self.scope_id,  # tenancy boundary
                col(Payment.deleted_at).is_(None),
                col(Payment.voided_at).is_(None),
                col(Payment.paid_at) >= lower,
                col(Payment.paid_at) < upper,
            )
            .group_by(col(Payment.method))
        )
        out = [
            MethodBreakdown(
                method=PaymentMethod(method), total=to_money(total), count=int(count or 0)
            )
            for method, total, count in self.session.execute(stmt).all()
        ]
        out.sort(key=lambda m: m.total, reverse=True)
        return out

    def _by_status(self, local_start: date, local_end: date) -> list[StatusBreakdown]:
        stmt = (
            select(
                col(Credit.status),
                func.count().label("count"),
                money_sum(Credit.grand_total).label("total"),
                money_sum(Credit.remaining_amount).label("outstanding"),
            )
            .where(
                Credit.business_id == self.scope_id,  # tenancy boundary
                col(Credit.deleted_at).is_(None),
                col(Credit.issued_date) >= local_start,
                col(Credit.issued_date) <= local_end,
            )
            .group_by(col(Credit.status))
        )
        return [
            StatusBreakdown(
                status=CreditStatus(status),
                count=int(count or 0),
                total=to_money(total),
                outstanding=to_money(outstanding),
            )
            for status, count, total, outstanding in self.session.execute(stmt).all()
        ]

    # ================================================================== export
    # All three renderers return BYTES. Nothing is written to disk -- see the module
    # docstring.
    def to_csv(self, report: ReportData) -> bytes:
        buf = io.StringIO()
        w = csv.writer(buf)
        money = report.currency

        w.writerow([report.title])
        w.writerow(["Business", report.business_name])
        w.writerow(["Generated", ensure_utc(report.generated_at).isoformat()])
        w.writerow(["Timezone", report.timezone])
        w.writerow([])

        s = report.summary
        w.writerow(["SUMMARY"])
        w.writerow(["Metric", f"Value ({money})"])
        w.writerow(["Credits issued", f"{s.credits_issued:.2f}"])
        w.writerow(["Credit count", s.credit_count])
        w.writerow(["Collected", f"{s.collected:.2f}"])
        w.writerow(["Payment count", s.payment_count])
        w.writerow(["Outstanding at period end", f"{s.outstanding_at_end:.2f}"])
        w.writerow(["Overdue amount", f"{s.overdue_amount:.2f}"])
        w.writerow(["Overdue count", s.overdue_count])
        w.writerow(["Collection rate %", f"{s.collection_rate_pct:.2f}"])
        w.writerow([])

        w.writerow(["BREAKDOWN"])
        w.writerow(
            [
                "Period" if report.granularity == "month" else "Date",
                f"Credits issued ({money})",
                "Credits",
                f"Collected ({money})",
                "Payments",
                f"Net ({money})",
            ]
        )
        for r in report.rows:
            w.writerow(
                [
                    r.label,
                    f"{r.credits_issued:.2f}",
                    r.credit_count,
                    f"{r.collected:.2f}",
                    r.payment_count,
                    f"{r.net:.2f}",
                ]
            )
        w.writerow([])

        w.writerow(["TOP CUSTOMERS"])
        w.writerow(["Code", "Name", f"Issued ({money})", "Credits", f"Collected ({money})", f"Outstanding ({money})"])
        for c in report.top_customers:
            w.writerow(
                [
                    c.code,
                    c.name,
                    f"{c.credits_issued:.2f}",
                    c.credit_count,
                    f"{c.collected:.2f}",
                    f"{c.outstanding:.2f}",
                ]
            )
        w.writerow([])

        w.writerow(["BY PAYMENT METHOD"])
        w.writerow(["Method", f"Total ({money})", "Payments"])
        for m in report.by_method:
            w.writerow([m.method.value, f"{m.total:.2f}", m.count])
        w.writerow([])

        w.writerow(["BY STATUS"])
        w.writerow(["Status", "Credits", f"Total ({money})", f"Outstanding ({money})"])
        for st in report.by_status:
            w.writerow([st.status.value, st.count, f"{st.total:.2f}", f"{st.outstanding:.2f}"])

        # utf-8-sig: Excel on Windows renders a plain UTF-8 CSV as mojibake without a
        # BOM, and a shopkeeper opening a garbled name column will not trust the file.
        return buf.getvalue().encode("utf-8-sig")

    def to_xlsx(self, report: ReportData) -> bytes:
        wb = Workbook()
        fmt = _money_format(report.currency_symbol)

        summary = wb.active
        assert summary is not None
        summary.title = "Summary"
        _xlsx_title(summary, report)

        s = report.summary
        _xlsx_header(summary, 6, ["Metric", "Value"])
        rows: list[tuple[str, Any, bool]] = [
            ("Credits issued", s.credits_issued, True),
            ("Credit count", s.credit_count, False),
            ("Collected", s.collected, True),
            ("Payment count", s.payment_count, False),
            ("Outstanding at period end", s.outstanding_at_end, True),
            ("Overdue amount", s.overdue_amount, True),
            ("Overdue count", s.overdue_count, False),
            ("Collection rate %", s.collection_rate_pct, False),
        ]
        for i, (label, value, is_money) in enumerate(rows, start=7):
            summary.cell(row=i, column=1, value=label)
            cell = summary.cell(row=i, column=2, value=_num(value))
            if is_money:
                cell.number_format = fmt
        _finish_sheet(summary, freeze="A7")

        breakdown = wb.create_sheet("Breakdown")
        _xlsx_header(
            breakdown,
            1,
            [
                "Period" if report.granularity == "month" else "Date",
                "Credits issued",
                "Credits",
                "Collected",
                "Payments",
                "Net",
            ],
        )
        for i, r in enumerate(report.rows, start=2):
            breakdown.cell(row=i, column=1, value=r.label)
            breakdown.cell(row=i, column=2, value=_num(r.credits_issued)).number_format = fmt
            breakdown.cell(row=i, column=3, value=r.credit_count)
            breakdown.cell(row=i, column=4, value=_num(r.collected)).number_format = fmt
            breakdown.cell(row=i, column=5, value=r.payment_count)
            breakdown.cell(row=i, column=6, value=_num(r.net)).number_format = fmt
        _finish_sheet(breakdown)

        customers = wb.create_sheet("Top customers")
        _xlsx_header(
            customers, 1, ["Code", "Customer", "Issued", "Credits", "Collected", "Outstanding"]
        )
        for i, c in enumerate(report.top_customers, start=2):
            customers.cell(row=i, column=1, value=c.code)
            customers.cell(row=i, column=2, value=c.name)
            customers.cell(row=i, column=3, value=_num(c.credits_issued)).number_format = fmt
            customers.cell(row=i, column=4, value=c.credit_count)
            customers.cell(row=i, column=5, value=_num(c.collected)).number_format = fmt
            customers.cell(row=i, column=6, value=_num(c.outstanding)).number_format = fmt
        _finish_sheet(customers)

        methods = wb.create_sheet("By method")
        _xlsx_header(methods, 1, ["Method", "Total", "Payments"])
        for i, m in enumerate(report.by_method, start=2):
            methods.cell(row=i, column=1, value=m.method.value)
            methods.cell(row=i, column=2, value=_num(m.total)).number_format = fmt
            methods.cell(row=i, column=3, value=m.count)
        _finish_sheet(methods)

        statuses = wb.create_sheet("By status")
        _xlsx_header(statuses, 1, ["Status", "Credits", "Total", "Outstanding"])
        for i, st in enumerate(report.by_status, start=2):
            statuses.cell(row=i, column=1, value=st.status.value)
            statuses.cell(row=i, column=2, value=st.count)
            statuses.cell(row=i, column=3, value=_num(st.total)).number_format = fmt
            statuses.cell(row=i, column=4, value=_num(st.outstanding)).number_format = fmt
        _finish_sheet(statuses)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def to_pdf(self, report: ReportData) -> bytes:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
            title=report.title,
            author=report.business_name,
        )
        st = _styles()
        sym = report.currency_symbol
        flow: list[Any] = [
            Paragraph(report.business_name, st["DocTitle"]),
            Paragraph(report.title, st["DocSubtitle"]),
            Paragraph(
                f"Generated {ensure_utc(report.generated_at):%d %b %Y %H:%M} UTC "
                f"&middot; times shown in {report.timezone}",
                st["Muted"],
            ),
            Spacer(1, 8 * mm),
        ]

        s = report.summary
        flow.append(
            _kpi_grid(
                [
                    ("Credits issued", _money(s.credits_issued, sym), f"{s.credit_count} records"),
                    ("Collected", _money(s.collected, sym), f"{s.payment_count} payments"),
                    ("Outstanding", _money(s.outstanding_at_end, sym), "at period end"),
                    ("Overdue", _money(s.overdue_amount, sym), f"{s.overdue_count} credits"),
                ]
            )
        )
        flow.append(Spacer(1, 8 * mm))

        if report.rows:
            flow.append(Paragraph("Breakdown", st["Section"]))
            head = "Period" if report.granularity == "month" else "Date"
            data = [[head, "Issued", "Credits", "Collected", "Payments", "Net"]] + [
                [
                    r.label,
                    _money(r.credits_issued, sym),
                    str(r.credit_count),
                    _money(r.collected, sym),
                    str(r.payment_count),
                    _money(r.net, sym),
                ]
                for r in report.rows
            ]
            flow.append(_data_table(data, numeric_from=1))
            flow.append(Spacer(1, 6 * mm))

        if report.top_customers:
            flow.append(Paragraph("Top customers", st["Section"]))
            data = [["Code", "Customer", "Issued", "Collected", "Outstanding"]] + [
                [
                    c.code,
                    c.name,
                    _money(c.credits_issued, sym),
                    _money(c.collected, sym),
                    _money(c.outstanding, sym),
                ]
                for c in report.top_customers
            ]
            flow.append(_data_table(data, numeric_from=2))
            flow.append(Spacer(1, 6 * mm))

        if report.by_method:
            flow.append(Paragraph("Collections by method", st["Section"]))
            data = [["Method", "Total", "Payments"]] + [
                [m.method.value.replace("_", " ").title(), _money(m.total, sym), str(m.count)]
                for m in report.by_method
            ]
            flow.append(_data_table(data, numeric_from=1))
            flow.append(Spacer(1, 6 * mm))

        if report.by_status:
            flow.append(Paragraph("Credits by status", st["Section"]))
            data = [["Status", "Credits", "Total", "Outstanding"]] + [
                [
                    b.status.value.replace("_", " ").title(),
                    str(b.count),
                    _money(b.total, sym),
                    _money(b.outstanding, sym),
                ]
                for b in report.by_status
            ]
            flow.append(_data_table(data, numeric_from=1))

        doc.build(flow, onFirstPage=_page_footer, onLaterPages=_page_footer)
        return buf.getvalue()

    # ================================================ invoice & receipt (A4)
    # Returned as bytes and NEVER persisted. See the module docstring: a generated
    # document is a view of the ledger, not a new record. No FileAsset, no upload.
    async def invoice_pdf(self, credit_id: str) -> bytes:
        self.require(Permission.REPORT_READ)
        business = self.get_business()
        credit = self._get_credit(credit_id)
        customer = self._get_customer(credit.customer_id)
        items = self._items(credit.id)
        payments = self._payments(credit.id)
        logo = await self._logo_bytes(business)

        sym = business.currency_symbol
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=15 * mm,
            bottomMargin=20 * mm,
            title=f"Invoice {credit.number}",
            author=business.name,
        )
        st = _styles()
        flow: list[Any] = [
            _letterhead(business, logo, "INVOICE", credit.number, st),
            Spacer(1, 8 * mm),
            _parties_block(
                business,
                customer,
                st,
                [
                    ("Invoice no.", credit.number),
                    ("Issued", f"{credit.issued_date:%d %b %Y}"),
                    ("Due", f"{credit.due_date:%d %b %Y}"),
                    ("Status", CreditStatus(credit.status).value.replace("_", " ").title()),
                ],
            ),
            Spacer(1, 7 * mm),
        ]

        # --- line items -----------------------------------------------------
        data: list[list[Any]] = [["#", "Description", "Qty", "Unit price", "Discount", "Tax", "Amount"]]
        for n, it in enumerate(items, start=1):
            desc = it.name if not it.description else f"{it.name}<br/><font size=7 color='#6B7280'>{_esc(it.description)}</font>"
            data.append(
                [
                    str(n),
                    Paragraph(desc, st["Cell"]),
                    f"{_trim(it.quantity)} {it.unit}",
                    _money(it.unit_price, sym),
                    _money(it.discount_amount, sym),
                    _money(it.tax_amount, sym),
                    _money(it.line_total, sym),
                ]
            )
        flow.append(_items_table(data))
        flow.append(Spacer(1, 5 * mm))

        # --- totals ---------------------------------------------------------
        totals: list[tuple[str, str, bool]] = [
            ("Subtotal", _money(credit.subtotal, sym), False),
        ]
        if credit.discount_amount > ZERO:
            totals.append(("Discount", f"-{_money(credit.discount_amount, sym)}", False))
        if credit.tax_amount > ZERO:
            totals.append(("Tax", _money(credit.tax_amount, sym), False))
        totals.append(("Total", _money(credit.grand_total, sym), True))
        totals.append(("Amount paid", _money(credit.amount_paid, sym), False))
        totals.append(("Balance due", _money(credit.remaining_amount, sym), True))
        flow.append(_totals_block(totals, st))

        if payments:
            flow.append(Spacer(1, 7 * mm))
            flow.append(Paragraph("Payments received", st["Section"]))
            pdata = [["Date", "Receipt", "Method", "Amount"]] + [
                [
                    f"{ensure_utc(p.paid_at).astimezone(get_tz(business.timezone)):%d %b %Y}",
                    p.number,
                    PaymentMethod(p.method).value.replace("_", " ").title(),
                    _money(p.amount, sym),
                ]
                for p in payments
            ]
            flow.append(_data_table(pdata, numeric_from=3))

        if credit.notes:
            flow.append(Spacer(1, 6 * mm))
            flow.append(Paragraph("Notes", st["Section"]))
            flow.append(Paragraph(_esc(credit.notes), st["Body"]))

        flow.append(Spacer(1, 8 * mm))
        flow.append(Paragraph(_footer_text(business, credit.due_date), st["Muted"]))

        doc.build(flow, onFirstPage=_page_footer, onLaterPages=_page_footer)
        return buf.getvalue()

    async def receipt_pdf(self, payment_id: str) -> bytes:
        self.require(Permission.REPORT_READ)
        business = self.get_business()
        payment = self._get_payment(payment_id)
        credit = self._get_credit(payment.credit_id)
        customer = self._get_customer(payment.customer_id)
        logo = await self._logo_bytes(business)

        sym = business.currency_symbol
        tz = get_tz(business.timezone)
        local_paid = ensure_utc(payment.paid_at).astimezone(tz)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=15 * mm,
            bottomMargin=20 * mm,
            title=f"Receipt {payment.number}",
            author=business.name,
        )
        st = _styles()
        flow: list[Any] = [
            _letterhead(business, logo, "RECEIPT", payment.number, st),
            Spacer(1, 8 * mm),
            _parties_block(
                business,
                customer,
                st,
                [
                    ("Receipt no.", payment.number),
                    ("Date", f"{local_paid:%d %b %Y %H:%M}"),
                    ("Against", credit.number),
                    ("Method", PaymentMethod(payment.method).value.replace("_", " ").title()),
                ],
            ),
            Spacer(1, 7 * mm),
        ]

        if payment.voided_at:
            # A voided receipt must SAY so. Handing a customer a receipt for a payment
            # that has been reversed is how disputes start.
            flow.append(Paragraph("*** VOIDED *** " + _esc(payment.void_reason or ""), st["Void"]))
            flow.append(Spacer(1, 4 * mm))

        rows: list[list[Any]] = [
            ["Description", "Amount"],
            [Paragraph(f"Payment against credit {credit.number}", st["Cell"]), _money(payment.amount, sym)],
        ]
        if payment.reference:
            rows.append([Paragraph(f"Reference: {_esc(payment.reference)}", st["Cell"]), ""])
        flow.append(_items_table(rows, widths=[120 * mm, 54 * mm]))
        flow.append(Spacer(1, 5 * mm))

        flow.append(
            _totals_block(
                [
                    ("Credit total", _money(credit.grand_total, sym), False),
                    ("Amount received", _money(payment.amount, sym), True),
                    ("Balance after payment", _money(payment.balance_after, sym), True),
                ],
                st,
            )
        )

        if payment.notes:
            flow.append(Spacer(1, 6 * mm))
            flow.append(Paragraph("Notes", st["Section"]))
            flow.append(Paragraph(_esc(payment.notes), st["Body"]))

        flow.append(Spacer(1, 10 * mm))
        flow.append(
            Paragraph(
                f"Received with thanks by {_esc(business.name)}. "
                f"This receipt confirms a payment of {_money(payment.amount, sym)} "
                f"on {local_paid:%d %b %Y}.",
                st["Muted"],
            )
        )
        doc.build(flow, onFirstPage=_page_footer, onLaterPages=_page_footer)
        return buf.getvalue()

    # ---------------------------------------------------------------- helpers
    def _get_credit(self, credit_id: str) -> Credit:
        credit = self.session.get(Credit, credit_id)
        if credit is None or credit.deleted_at is not None:
            raise NotFoundError("Credit record not found")
        self.assert_in_scope(credit.business_id)
        return credit

    def _get_payment(self, payment_id: str) -> Payment:
        payment = self.session.get(Payment, payment_id)
        if payment is None or payment.deleted_at is not None:
            raise NotFoundError("Payment not found")
        self.assert_in_scope(payment.business_id)
        return payment

    def _get_customer(self, customer_id: str) -> Customer:
        customer = self.session.get(Customer, customer_id)
        if customer is None:
            raise NotFoundError("Customer not found")
        self.assert_in_scope(customer.business_id)
        return customer

    def _items(self, credit_id: str) -> list[CreditItem]:
        stmt = (
            select(CreditItem)
            .where(
                CreditItem.business_id == self.scope_id,  # tenancy boundary
                CreditItem.credit_id == credit_id,
            )
            .order_by(col(CreditItem.position).asc())
        )
        return list(self.session.execute(stmt).scalars().all())

    def _payments(self, credit_id: str) -> list[Payment]:
        stmt = (
            select(Payment)
            .where(
                Payment.business_id == self.scope_id,  # tenancy boundary
                Payment.credit_id == credit_id,
                col(Payment.deleted_at).is_(None),
                col(Payment.voided_at).is_(None),
            )
            .order_by(col(Payment.paid_at).asc())
        )
        return list(self.session.execute(stmt).scalars().all())

    async def _logo_bytes(self, business: Business) -> bytes | None:
        """Fetch the logo through StorageService. A missing logo is normal, not an error."""
        if not business.logo_file_id:
            return None
        asset = self.session.get(FileAsset, business.logo_file_id)
        if asset is None or asset.business_id != business.id or asset.kind != FileKind.BUSINESS_LOGO:
            return None
        try:
            return await StorageService(self.session).read(asset)
        except (FileNotFoundError, OSError):
            # The FileAsset row outlived its bytes (a half-finished restore, an S3
            # blip). Print the invoice without a logo rather than fail the download --
            # the customer is standing at the counter.
            return None


# ---------------------------------------------------------------------------
# PDF furniture
# ---------------------------------------------------------------------------
_INK = colors.HexColor("#111827")
_MUTED = colors.HexColor("#6B7280")
_LINE = colors.HexColor("#E5E7EB")
_BAND = colors.HexColor("#F3F4F6")
_ACCENT = colors.HexColor("#4F46E5")


def _styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "DocTitle": ParagraphStyle(
            "DocTitle", parent=base["Title"], fontSize=18, leading=22, alignment=0, textColor=_INK
        ),
        "DocSubtitle": ParagraphStyle(
            "DocSubtitle", parent=base["Normal"], fontSize=11, leading=14, textColor=_MUTED
        ),
        "DocType": ParagraphStyle(
            "DocType",
            parent=base["Normal"],
            fontSize=20,
            leading=24,
            alignment=TA_RIGHT,
            textColor=_ACCENT,
            fontName="Helvetica-Bold",
        ),
        "DocNumber": ParagraphStyle(
            "DocNumber", parent=base["Normal"], fontSize=10, alignment=TA_RIGHT, textColor=_MUTED
        ),
        "Section": ParagraphStyle(
            "Section",
            parent=base["Normal"],
            fontSize=11,
            leading=14,
            spaceAfter=4,
            textColor=_INK,
            fontName="Helvetica-Bold",
        ),
        "Body": ParagraphStyle("Body", parent=base["Normal"], fontSize=9, leading=12, textColor=_INK),
        "Cell": ParagraphStyle("Cell", parent=base["Normal"], fontSize=8.5, leading=11, textColor=_INK),
        "Muted": ParagraphStyle("Muted", parent=base["Normal"], fontSize=8, leading=11, textColor=_MUTED),
        "Void": ParagraphStyle(
            "Void",
            parent=base["Normal"],
            fontSize=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#B91C1C"),
            fontName="Helvetica-Bold",
        ),
        "KpiLabel": ParagraphStyle("KpiLabel", parent=base["Normal"], fontSize=8, textColor=_MUTED),
        "KpiValue": ParagraphStyle(
            "KpiValue", parent=base["Normal"], fontSize=13, leading=16, textColor=_INK, fontName="Helvetica-Bold"
        ),
    }


def _letterhead(
    business: Business, logo: bytes | None, doc_type: str, number: str, st: dict[str, ParagraphStyle]
) -> Table:
    left: list[Any] = []
    if logo:
        try:
            img = Image(io.BytesIO(logo))
            # Fit inside a 38x18mm box, preserving aspect. A stretched logo looks
            # amateurish on a document the shop hands to a customer.
            ratio = (img.imageWidth or 1) / (img.imageHeight or 1)
            height = min(18 * mm, 38 * mm / ratio if ratio else 18 * mm)
            img.drawHeight = height
            img.drawWidth = height * ratio
            img.hAlign = "LEFT"
            left.append(img)
            left.append(Spacer(1, 3 * mm))
        except Exception:  # noqa: BLE001 - a bad logo must never break the invoice
            pass

    left.append(Paragraph(_esc(business.name), st["DocTitle"]))
    for line in _business_lines(business):
        left.append(Paragraph(line, st["Muted"]))

    right = [Paragraph(doc_type, st["DocType"]), Paragraph(_esc(number), st["DocNumber"])]

    table = Table([[left, right]], colWidths=[110 * mm, 64 * mm])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("LINEBELOW", (0, 0), (-1, -1), 1, _ACCENT),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _business_lines(business: Business) -> list[str]:
    parts = [
        ", ".join(p for p in (business.address, business.city, business.country) if p),
        " | ".join(p for p in (business.phone, business.email) if p),
        business.website or "",
    ]
    return [_esc(p) for p in parts if p]


def _parties_block(
    business: Business,
    customer: Customer,
    st: dict[str, ParagraphStyle],
    meta: list[tuple[str, str]],
) -> Table:
    bill_to = [Paragraph("BILL TO", st["Section"]), Paragraph(_esc(customer.name), st["Body"])]
    for line in (
        customer.code,
        customer.phone,
        customer.email,
        ", ".join(p for p in (customer.address, customer.city) if p),
    ):
        if line:
            bill_to.append(Paragraph(_esc(line), st["Muted"]))

    meta_rows = [[Paragraph(f"<b>{_esc(k)}</b>", st["Muted"]), Paragraph(_esc(v), st["Body"])] for k, v in meta]
    meta_table = Table(meta_rows, colWidths=[26 * mm, 38 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    outer = Table([[bill_to, meta_table]], colWidths=[110 * mm, 64 * mm])
    outer.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return outer


def _items_table(data: list[list[Any]], widths: list[float] | None = None) -> Table:
    cols = widths or [8 * mm, 62 * mm, 20 * mm, 22 * mm, 20 * mm, 20 * mm, 22 * mm]
    table = Table(data, colWidths=cols, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _INK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _BAND]),
                ("LINEBELOW", (0, 1), (-1, -1), 0.4, _LINE),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _totals_block(rows: list[tuple[str, str, bool]], st: dict[str, ParagraphStyle]) -> KeepTogether:
    """Right-aligned totals column. Bold rows are the ones the customer actually reads."""
    data = [
        [
            Paragraph(f"<b>{_esc(label)}</b>" if strong else _esc(label), st["Body"]),
            Paragraph(f"<b>{_esc(value)}</b>" if strong else _esc(value), st["Body"]),
        ]
        for label, value, strong in rows
    ]
    inner = Table(data, colWidths=[38 * mm, 32 * mm], hAlign="RIGHT")
    style = [
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEABOVE", (0, 0), (-1, 0), 0.6, _LINE),
    ]
    for i, (_, _, strong) in enumerate(rows):
        if strong:
            style.append(("LINEABOVE", (0, i), (-1, i), 0.8, _INK))
            style.append(("BACKGROUND", (0, i), (-1, i), _BAND))
    inner.setStyle(TableStyle(style))
    return KeepTogether(inner)


def _data_table(data: list[list[Any]], numeric_from: int = 1) -> Table:
    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _BAND),
                ("TEXTCOLOR", (0, 0), (-1, 0), _INK),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (numeric_from, 0), (-1, -1), "RIGHT"),
                ("LINEBELOW", (0, 0), (-1, -1), 0.3, _LINE),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _kpi_grid(cards: list[tuple[str, str, str]]) -> Table:
    st = _styles()
    cells = [
        [
            Paragraph(_esc(label), st["KpiLabel"]),
            Paragraph(_esc(value), st["KpiValue"]),
            Paragraph(_esc(note), st["KpiLabel"]),
        ]
        for label, value, note in cards
    ]
    table = Table([cells], colWidths=[43.5 * mm] * len(cells))
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOX", (0, 0), (-1, -1), 0.5, _LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, _LINE),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _page_footer(canvas: Any, doc: Any) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(_MUTED)
    canvas.drawString(18 * mm, 12 * mm, f"Generated {utcnow():%d %b %Y %H:%M} UTC")
    canvas.drawRightString(A4[0] - 18 * mm, 12 * mm, f"Page {canvas.getPageNumber()}")
    canvas.setStrokeColor(_LINE)
    canvas.line(18 * mm, 15 * mm, A4[0] - 18 * mm, 15 * mm)
    canvas.restoreState()


def _footer_text(business: Business, due: date) -> str:
    bits = [f"Payment is due by {due:%d %b %Y}."]
    if business.phone:
        bits.append(f"Questions? Call {_esc(business.phone)}.")
    bits.append("Thank you for your business.")
    return " ".join(bits)


def _money(value: Decimal, symbol: str) -> str:
    """Render an amount with the business's own currency symbol.

    The space is conditional because the convention differs by symbol shape: a
    glyph binds tight ("$1,234.50", "€1,234.50"), while a lettered or dotted
    abbreviation does not ("Nu. 1,234.50", "Rs. 1,234.50" -- "Nu.1,234.50" reads
    like a typo). Anything ending in a letter or a full stop gets the space.
    """
    amount = f"{quantize_money(value):,.2f}"
    gap = " " if symbol and (symbol[-1].isalpha() or symbol.endswith(".")) else ""
    return f"{symbol}{gap}{amount}"


def _trim(value: Decimal) -> str:
    """3 kg, not 3.000 kg."""
    q = Decimal(value).normalize()
    return f"{q:f}"


def _esc(text: str) -> str:
    """Escape for reportlab's mini-HTML. Customer names contain '&' more often than
    you would like, and an unescaped one aborts the whole PDF build."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# ---------------------------------------------------------------------------
# XLSX furniture
# ---------------------------------------------------------------------------
_XLSX_HEADER_FILL = PatternFill("solid", fgColor="111827")
_XLSX_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_XLSX_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


def _money_format(symbol: str) -> str:
    """Excel number format with the business's currency symbol.

    The symbol is quoted so a multi-character one (e.g. "Nu.") is not parsed as
    format tokens -- '.' and 'u' mean something to Excel.
    """
    safe = symbol.replace('"', "")
    return f'"{safe}"#,##0.00'


def _num(value: Any) -> Any:
    """Write money as a NUMBER, not a string.

    A Decimal is handed to openpyxl as a float here on purpose: Excel has no decimal
    type, and a string would make the column unsummable in the very spreadsheet the
    user opened it in to do exactly that. The stored value remains exact; this is a
    presentation copy.
    """
    return float(value) if isinstance(value, Decimal) else value


def _xlsx_title(sheet: Worksheet, report: ReportData) -> None:
    sheet["A1"] = report.business_name
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = report.title
    sheet["A3"] = f"Generated {ensure_utc(report.generated_at):%Y-%m-%d %H:%M} UTC ({report.timezone})"
    sheet["A3"].font = Font(size=9, color="6B7280")
    sheet["A4"] = f"Currency: {report.currency}"
    sheet["A4"].font = Font(size=9, color="6B7280")


def _xlsx_header(sheet: Worksheet, row: int, headers: list[str]) -> None:
    for i, name in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=i, value=name)
        cell.fill = _XLSX_HEADER_FILL
        cell.font = _XLSX_HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _XLSX_BORDER
    sheet.row_dimensions[row].height = 20


def _finish_sheet(sheet: Worksheet, freeze: str = "A2") -> None:
    """Freeze the header and auto-size the columns."""
    sheet.freeze_panes = freeze
    for column in sheet.columns:
        cells = list(column)
        width = max((len(str(c.value)) for c in cells if c.value is not None), default=8)
        letter = get_column_letter(cells[0].column or 1)
        sheet.column_dimensions[letter].width = min(50, max(11, width + 3))


def _bucket_end(bucket: date, granularity: Granularity) -> date:
    """Exclusive upper edge of a bucket."""
    if granularity == "day":
        return bucket + timedelta(days=1)
    return (
        bucket.replace(year=bucket.year + 1, month=1)
        if bucket.month == 12
        else bucket.replace(month=bucket.month + 1)
    )


__all__ = [
    "MethodBreakdown",
    "ReportCustomer",
    "ReportData",
    "ReportRow",
    "ReportService",
    "ReportSummary",
    "StatusBreakdown",
]
