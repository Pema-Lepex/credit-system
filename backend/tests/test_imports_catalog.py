"""Bulk import for products and services.

Same rules as the customer/credit importer (see test_imports.py): validate the whole
sheet, then write it all or write nothing. What is specific here is UNIQUENESS --
SKUs and codes must be unique per shop, including against other rows of the same
file, which the database cannot see because they are all still in flight.
"""

from __future__ import annotations

from decimal import Decimal

import pytest
from sqlmodel import select

from app.core.errors import ValidationError
from app.models.catalog import Category, Product, Service
from app.services.base import ServiceContext
from app.services.imports import ImportService

D = Decimal

PRODUCT_HEADER = "name,price,unit,category,sku,barcode,cost_price,stock_quantity,tax_percentage,is_active,description"
SERVICE_HEADER = "name,price,category,code,duration_minutes,tax_percentage,is_active,description"


def _csv(*lines: str) -> bytes:
    return "\n".join(lines).encode("utf-8")


def _run(ctx: ServiceContext, dataset: str, data: bytes, *, dry_run: bool = False):
    return ImportService(ctx).run(
        ctx, dataset=dataset, filename="sheet.csv", data=data, dry_run=dry_run
    )


def _products(session) -> list[Product]:
    return list(session.exec(select(Product).order_by(Product.name)).all())


def _services(session) -> list[Service]:
    return list(session.exec(select(Service).order_by(Service.name)).all())


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------
def test_imports_products(ctx, session):
    report = _run(
        ctx,
        "products",
        _csv(
            PRODUCT_HEADER,
            "Rice 5kg,450.00,kg,Groceries,RICE-5KG,8901234567890,380.00,24,0,YES,Local red rice",
            "Cooking oil 1L,220,litre,Groceries,OIL-1L,,180,12,,YES,",
            "Cigarettes,30,pcs,Tobacco,,,25,100,,,",
        ),
    )

    assert report.ok, report.errors
    assert report.created == 3

    items = _products(session)
    assert [p.name for p in items] == ["Cigarettes", "Cooking oil 1L", "Rice 5kg"]

    rice = items[2]
    assert rice.price == D("450.00")
    assert rice.cost_price == D("380.00")
    assert rice.unit == "kg"
    assert rice.sku == "RICE-5KG"
    assert rice.barcode == "8901234567890"
    assert rice.stock_quantity == D("24")
    assert rice.is_active is True
    assert rice.description == "Local red rice"


def test_only_the_name_is_required(ctx, session):
    report = _run(ctx, "products", _csv("name", "Cigarettes"))

    assert report.ok, report.errors
    product = _products(session)[0]
    assert product.price == D("0.00")
    assert product.unit == "pcs"  # the default, not blank
    assert product.is_active is True
    assert product.sku is None


def test_a_product_with_no_name_is_refused(ctx, session):
    report = _run(ctx, "products", _csv(PRODUCT_HEADER, ",450,kg,,,,,,,,"))

    assert not report.ok
    assert report.errors[0].column == "name"
    assert _products(session) == []


# ---------------------------------------------------------------------------
# Categories: created on demand, unlike customers
# ---------------------------------------------------------------------------
def test_categories_are_created_from_their_names(ctx, session):
    """A shopkeeper types "Groceries", not a UUID."""
    report = _run(
        ctx,
        "products",
        _csv("name,category", "Rice 5kg,Groceries", "Cigarettes,Tobacco"),
    )

    assert report.ok, report.errors
    names = sorted(c.name for c in session.exec(select(Category)).all())
    assert names == ["Groceries", "Tobacco"]

    rice = next(p for p in _products(session) if p.name == "Rice 5kg")
    groceries = next(c for c in session.exec(select(Category)).all() if c.name == "Groceries")
    assert rice.category_id == groceries.id


def test_one_category_is_created_once_however_many_rows_use_it(ctx, session):
    report = _run(
        ctx,
        "products",
        _csv(
            "name,category",
            "Rice 5kg,Groceries",
            "Cooking oil,Groceries",
            "Sugar,groceries",  # different case -- still the same shelf
        ),
    )

    assert report.ok, report.errors
    assert len(session.exec(select(Category)).all()) == 1
    ids = {p.category_id for p in _products(session)}
    assert len(ids) == 1


def test_an_existing_category_is_reused_not_duplicated(ctx, session, business):
    session.add(Category(business_id=business.id, name="Groceries"))
    session.commit()

    report = _run(ctx, "products", _csv("name,category", "Rice 5kg,GROCERIES"))

    assert report.ok, report.errors
    assert len(session.exec(select(Category)).all()) == 1


def test_a_dry_run_creates_no_categories(ctx, session):
    """THE purity rule. Validators must not write -- a preview that creates
    categories has written to a database the caller was promised was untouched."""
    report = _run(ctx, "products", _csv("name,category", "Rice 5kg,Groceries"), dry_run=True)

    assert report.ok
    assert report.created == 0
    assert session.exec(select(Category)).all() == []
    assert _products(session) == []


# ---------------------------------------------------------------------------
# Uniqueness — the part the database cannot check alone
# ---------------------------------------------------------------------------
def test_a_sku_already_in_the_shop_is_refused_by_name(ctx, session, business):
    session.add(Product(business_id=business.id, name="Old Rice", sku="RICE-5KG"))
    session.commit()

    report = _run(ctx, "products", _csv("name,sku", "Rice 5kg,RICE-5KG"))

    assert not report.ok
    assert report.errors[0].column == "sku"
    assert "Old Rice" in report.errors[0].message  # says WHICH product owns it
    assert len(_products(session)) == 1  # nothing new landed


def test_a_sku_used_twice_in_the_same_file_is_refused(ctx, session):
    """The database cannot see this: both rows are still in flight."""
    report = _run(
        ctx,
        "products",
        _csv("name,sku", "Rice 5kg,RICE-5KG", "Cooking oil,OIL-1L", "Red rice,RICE-5KG"),
    )

    assert not report.ok
    assert report.errors[0].column == "sku"
    assert "row 2" in report.errors[0].message  # points at the FIRST use
    assert _products(session) == []


def test_a_duplicate_barcode_in_the_file_is_refused(ctx, session):
    report = _run(
        ctx, "products", _csv("name,barcode", "Rice,890123", "Oil,890123")
    )
    assert not report.ok
    assert report.errors[0].column == "barcode"


def test_blank_skus_do_not_collide(ctx, session):
    """Most shops do not use SKUs at all. Empty is not a duplicate of empty."""
    report = _run(ctx, "products", _csv("name,sku", "Rice 5kg,", "Cooking oil,", "Sugar,"))

    assert report.ok, report.errors
    assert report.created == 3


# ---------------------------------------------------------------------------
# Field parsing
# ---------------------------------------------------------------------------
def test_yes_no_accepts_what_real_sheets_contain(ctx, session):
    report = _run(
        ctx,
        "products",
        _csv(
            "name,is_active",
            "A,YES",
            "B,no",
            "C,TRUE",
            "D,0",
            "E,",  # blank -> default YES
        ),
    )

    assert report.ok, report.errors
    active = {p.name: p.is_active for p in _products(session)}
    assert active == {"A": True, "B": False, "C": True, "D": False, "E": True}


def test_an_unclear_active_value_is_refused(ctx):
    report = _run(ctx, "products", _csv("name,is_active", "Rice,maybe"), dry_run=True)
    assert not report.ok
    assert report.errors[0].column == "is_active"


def test_zero_stock_is_allowed(ctx, session):
    """"I have none left" is a fact, not a typo."""
    report = _run(ctx, "products", _csv("name,stock_quantity", "Rice 5kg,0"))

    assert report.ok, report.errors
    assert _products(session)[0].stock_quantity == D("0")


def test_negative_stock_is_refused(ctx):
    report = _run(ctx, "products", _csv("name,stock_quantity", "Rice,-5"), dry_run=True)
    assert not report.ok
    assert report.errors[0].column == "stock_quantity"


def test_money_survives_currency_formatting(ctx, session):
    report = _run(ctx, "products", _csv("name,price", 'Rice 5kg,"Nu. 1,450.50"'))

    assert report.ok, report.errors
    assert _products(session)[0].price == D("1450.50")


def test_a_bad_tax_percentage_is_refused(ctx):
    report = _run(ctx, "products", _csv("name,tax_percentage", "Rice,150"), dry_run=True)
    assert not report.ok
    assert report.errors[0].column == "tax_percentage"


# ---------------------------------------------------------------------------
# Services
# ---------------------------------------------------------------------------
def test_imports_services(ctx, session):
    report = _run(
        ctx,
        "services",
        _csv(
            SERVICE_HEADER,
            "Puncture repair,150.00,Repairs,SVC-PUNC,30,0,YES,Includes tube patch",
            "Full service,1200,Repairs,SVC-FULL,120,,YES,",
        ),
    )

    assert report.ok, report.errors
    assert report.created == 2

    items = _services(session)
    puncture = next(s for s in items if s.name == "Puncture repair")
    assert puncture.price == D("150.00")
    assert puncture.code == "SVC-PUNC"
    assert puncture.duration_minutes == 30
    assert puncture.description == "Includes tube patch"

    # Both rows named Repairs -> one category.
    assert len(session.exec(select(Category)).all()) == 1


def test_a_duplicate_service_code_is_refused(ctx, session):
    report = _run(
        ctx, "services", _csv("name,code", "Puncture repair,SVC-1", "Tyre change,SVC-1")
    )
    assert not report.ok
    assert report.errors[0].column == "code"
    assert _services(session) == []


def test_a_bad_duration_is_refused(ctx):
    report = _run(
        ctx, "services", _csv("name,duration_minutes", "Repair,half an hour"), dry_run=True
    )
    assert not report.ok
    assert report.errors[0].column == "duration_minutes"


# ---------------------------------------------------------------------------
# The shared rules still hold
# ---------------------------------------------------------------------------
def test_one_bad_row_aborts_the_whole_product_batch(ctx, session):
    """R1, on the new datasets."""
    report = _run(
        ctx,
        "products",
        _csv("name,price", "Rice 5kg,450", ",100", "Sugar,60"),
    )

    assert not report.ok
    assert report.created == 0
    assert _products(session) == []
    assert session.exec(select(Category)).all() == []  # no categories either


def test_every_bad_product_row_is_reported_at_once(ctx):
    report = _run(
        ctx,
        "products",
        _csv(
            "name,price,stock_quantity",
            "Rice,notmoney,10",
            "Oil,100,-5",
            ",100,10",
        ),
        dry_run=True,
    )

    assert {(e.row, e.column) for e in report.errors} == {
        (2, "price"),
        (3, "stock_quantity"),
        (4, "name"),
    }


def test_a_dry_run_of_products_writes_nothing(ctx, session):
    report = _run(ctx, "products", _csv("name,price", "Rice 5kg,450"), dry_run=True)

    assert report.ok
    assert report.created == 0
    assert _products(session) == []


def test_human_labels_work_as_product_headings(ctx, session):
    report = _run(ctx, "products", _csv("Name,Selling price,SKU", "Rice 5kg,450,RICE-1"))

    assert report.ok, report.errors
    assert _products(session)[0].sku == "RICE-1"


def test_unknown_product_columns_are_ignored_with_a_warning(ctx, session):
    report = _run(ctx, "products", _csv("name,shelf_number", "Rice 5kg,B4"))

    assert report.ok, report.errors
    assert any("shelf_number" in w.message for w in report.warnings)


# ---------------------------------------------------------------------------
# Templates + tenancy
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("dataset", ["products", "services"])
def test_templates_are_generated_for_the_new_datasets(ctx, dataset: str):
    import io as _io

    from openpyxl import load_workbook

    data, filename, _ = ImportService(ctx).template(dataset, "xlsx")
    assert filename == f"{dataset}-import-template.xlsx"

    wb = load_workbook(_io.BytesIO(data))
    assert wb.sheetnames == [dataset.capitalize(), "Instructions"]
    sheet = wb[dataset.capitalize()]
    assert [c.value for c in sheet[1]][0] == "name"
    assert sheet.max_row == 1  # headings only


@pytest.mark.parametrize("dataset", ["products", "services"])
def test_csv_templates_are_headings_only(ctx, dataset: str):
    data, _, _ = ImportService(ctx).template(dataset, "csv")
    text = data.decode("utf-8-sig")
    assert len(text.strip().splitlines()) == 1
    assert text.startswith("name,")


def test_a_sku_belonging_to_another_shop_does_not_collide(ctx, session):
    """SKUs are unique PER BUSINESS -- two shops may both sell 'A-1'."""
    from app.models.business import Business

    other = Business(name="Rival Shop", slug="rival-shop", email="rival@x.bt")
    session.add(other)
    session.commit()
    session.add(Product(business_id=other.id, name="Their Rice", sku="RICE-5KG"))
    session.commit()

    report = _run(ctx, "products", _csv("name,sku", "Rice 5kg,RICE-5KG"))

    assert report.ok, report.errors
    assert report.created == 1


def test_an_unknown_dataset_is_still_rejected(ctx):
    with pytest.raises(ValidationError, match="Unknown import type"):
        _run(ctx, "widgets", _csv("name", "x"))
