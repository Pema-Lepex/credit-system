"""Bulk import: templates, validation, and the all-or-nothing rule.

The rule under test throughout is R1 from app/services/imports.py: a sheet is
validated in full and then written in full, or it is not written at all. Most of
these tests are really asking one question -- "did anything land when it
shouldn't have?"
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from sqlmodel import Session, select

from app.core.errors import ValidationError
from app.models.credit import Credit
from app.models.customer import Customer
from app.models.enums import CreditStatus, CustomerStatus
from app.services.base import ServiceContext
from app.services.imports import CUSTOMER_COLUMNS, ImportService

CUSTOMER_HEADER = "name,phone,email,city,status,credit_limit,notes"
CREDIT_HEADER = "customer_code,item_name,quantity,unit_price,due_date,issued_date,initial_payment"

# Credit status is derived against the real clock, so these are relative rather
# than literal -- a hardcoded date silently changes what the test asserts once it
# slips into the past.
TODAY = date.today()
_FUTURE_DAY = TODAY + timedelta(days=30)
_PAST_DAY = TODAY - timedelta(days=30)
FUTURE = _FUTURE_DAY.isoformat()
PAST = _PAST_DAY.isoformat()
# Credit numbers restart each year and are stamped from the ISSUED date, not today
# -- so in January the issued date is last year and so is the number.
ISSUED_YEAR = _PAST_DAY.year


def _csv(*lines: str) -> bytes:
    return "\n".join(lines).encode("utf-8")


def _run(ctx: ServiceContext, dataset: str, data: bytes, *, dry_run: bool = False):
    return ImportService(ctx).run(
        ctx, dataset=dataset, filename="sheet.csv", data=data, dry_run=dry_run
    )


def _customers(session: Session) -> list[Customer]:
    return list(session.exec(select(Customer).order_by(Customer.code)).all())


# ---------------------------------------------------------------------------
# Customers
# ---------------------------------------------------------------------------
def test_imports_customers_and_assigns_sequential_codes(ctx, session):
    report = _run(
        ctx,
        "customers",
        _csv(
            CUSTOMER_HEADER,
            "Sonam Dorji,+975 17 11 11 11,sonam@example.com,Thimphu,ACTIVE,5000,Regular",
            "Pema Lhamo,+975 17 22 22 22,,Paro,,,",
            "Karma Wangdi,,karma@example.com,,INACTIVE,250.50,",
        ),
    )

    assert report.ok, report.errors
    assert report.created == 3
    assert report.total_rows == 3

    people = _customers(session)
    assert [c.code for c in people] == ["CUST-0001", "CUST-0002", "CUST-0003"]
    assert [c.name for c in people] == ["Sonam Dorji", "Pema Lhamo", "Karma Wangdi"]

    sonam = people[0]
    assert sonam.email == "sonam@example.com"
    assert sonam.credit_limit == Decimal("5000.00")
    assert CustomerStatus(sonam.status) is CustomerStatus.ACTIVE
    # Blank status must fall back to ACTIVE rather than land as NULL/"".
    assert CustomerStatus(people[1].status) is CustomerStatus.ACTIVE
    assert CustomerStatus(people[2].status) is CustomerStatus.INACTIVE
    assert people[2].credit_limit == Decimal("250.50")


def test_dry_run_writes_nothing(ctx, session):
    report = _run(
        ctx,
        "customers",
        _csv(CUSTOMER_HEADER, "Sonam Dorji,,,,,,"),
        dry_run=True,
    )

    assert report.ok
    assert report.dry_run is True
    assert report.created == 0
    assert _customers(session) == []


def test_one_bad_row_aborts_the_whole_batch(ctx, session):
    """R1. The good rows either side of the bad one must NOT land."""
    report = _run(
        ctx,
        "customers",
        _csv(
            CUSTOMER_HEADER,
            "Sonam Dorji,,,,,,",
            ",,,,,,x",  # no name -- row 3
            "Karma Wangdi,,,,,,",
        ),
    )

    assert not report.ok
    assert report.created == 0
    assert _customers(session) == []

    assert [e.row for e in report.errors] == [3]
    assert report.errors[0].column == "name"


def test_every_bad_row_is_reported_at_once(ctx):
    """The preview exists to be exhaustive; one error per upload is useless."""
    report = _run(
        ctx,
        "customers",
        _csv(
            CUSTOMER_HEADER,
            "Sonam,,not-an-email,,,,",
            "Pema,,,,NOPE,,",
            "Karma,,,,,abc,",
        ),
        dry_run=True,
    )

    assert {(e.row, e.column) for e in report.errors} == {
        (2, "email"),
        (3, "status"),
        (4, "credit_limit"),
    }


def test_money_survives_currency_symbols_and_separators(ctx, session):
    report = _run(ctx, "customers", _csv(CUSTOMER_HEADER, 'Sonam,,,,,"Nu. 1,200.50",'))

    assert report.ok, report.errors
    assert _customers(session)[0].credit_limit == Decimal("1200.50")


def test_blank_rows_are_skipped_not_rejected(ctx, session):
    report = _run(
        ctx,
        "customers",
        _csv(CUSTOMER_HEADER, "Sonam,,,,,,", ",,,,,,", "", "Pema,,,,,,"),
    )

    assert report.ok, report.errors
    assert report.total_rows == 2
    assert report.created == 2


def test_duplicate_phone_warns_but_does_not_block(ctx, session, customer):
    report = _run(ctx, "customers", _csv(CUSTOMER_HEADER, "Someone Else,+975 17 12 34 56,,,,,"))

    assert report.ok
    assert report.created == 1
    assert len(report.warnings) == 1
    assert customer.code in report.warnings[0].message


def test_unknown_columns_are_ignored_with_a_warning(ctx, session):
    report = _run(ctx, "customers", _csv("name,my_own_notes", "Sonam,pays on fridays"))

    assert report.ok, report.errors
    assert report.created == 1
    assert any("my_own_notes" in w.message for w in report.warnings)


def test_human_labels_work_as_headings(ctx, session):
    """Someone will retype the headings by hand. 'Credit limit' must still land."""
    report = _run(ctx, "customers", _csv("Name,Credit limit", "Sonam Dorji,5000"))

    assert report.ok, report.errors
    assert _customers(session)[0].credit_limit == Decimal("5000.00")


def test_missing_required_column_is_rejected_up_front(ctx):
    with pytest.raises(ValidationError, match="name"):
        _run(ctx, "customers", _csv("phone,email", "+975 17,x@y.com"))


def test_semicolon_delimited_csv_is_understood(ctx, session):
    """European Excel writes these by default and they look identical to the owner."""
    report = _run(ctx, "customers", _csv("name;city", "Sonam Dorji;Thimphu"))

    assert report.ok, report.errors
    assert _customers(session)[0].city == "Thimphu"


# ---------------------------------------------------------------------------
# Credits
# ---------------------------------------------------------------------------
def test_imports_credits_against_an_existing_customer(ctx, session, customer):
    """Dates are relative to today on purpose.

    A credit's status is derived from its due date against the real clock, so a
    hardcoded 2026-08-01 quietly flips this test from PENDING to OVERDUE the day
    it passes -- and then it is asserting something nobody chose.
    """
    report = _run(
        ctx,
        "credits",
        _csv(
            CREDIT_HEADER,
            f"CUST-0001,Rice 5kg,2,450.00,{FUTURE},{PAST},",
            f"CUST-0001,Repair labour,1,1500,{FUTURE},{PAST},500",
        ),
    )

    assert report.ok, report.errors
    assert report.created == 2

    credits = list(session.exec(select(Credit).order_by(Credit.number)).all())
    assert [c.number for c in credits] == [
        f"CR-{ISSUED_YEAR}-0001",
        f"CR-{ISSUED_YEAR}-0002",
    ]

    rice = credits[0]
    assert rice.grand_total == Decimal("900.00")  # 2 x 450
    assert rice.remaining_amount == Decimal("900.00")
    assert CreditStatus(rice.status) is CreditStatus.PENDING

    # The initial payment must flow through PaymentService, not just set a column.
    repair = credits[1]
    assert repair.grand_total == Decimal("1500.00")
    assert repair.amount_paid == Decimal("500.00")
    assert repair.remaining_amount == Decimal("1000.00")
    assert CreditStatus(repair.status) is CreditStatus.PARTIALLY_PAID

    # R2: the customer's rolled-up totals were maintained by the real service.
    session.refresh(customer)
    assert customer.total_credit == Decimal("2400.00")
    assert customer.total_paid == Decimal("500.00")
    assert customer.outstanding_balance == Decimal("1900.00")
    assert customer.credit_count == 2


def test_a_past_due_date_imports_as_overdue(ctx, session, customer):
    """Importing history is the main use case, so most rows ARE already overdue."""
    report = _run(ctx, "credits", _csv(CREDIT_HEADER, f"CUST-0001,Rice,1,100,{PAST},{PAST},"))

    assert report.ok, report.errors
    assert CreditStatus(session.exec(select(Credit)).first().status) is CreditStatus.OVERDUE


def test_credit_matches_customer_by_phone(ctx, session, customer):
    report = _run(
        ctx,
        "credits",
        _csv(
            "customer_phone,item_name,unit_price,due_date",
            # Formatted differently to the stored number on purpose.
            f"9751712 3456,Cooking oil,220,{FUTURE}",
        ),
    )

    assert report.ok, report.errors
    assert session.exec(select(Credit)).first().customer_id == customer.id


def test_unknown_customer_is_rejected(ctx, session, customer):
    report = _run(
        ctx,
        "credits",
        _csv(CREDIT_HEADER, f"CUST-9999,Rice,1,100,{FUTURE},,"),
    )

    assert not report.ok
    assert report.created == 0
    assert report.errors[0].column == "customer_code"
    assert "CUST-9999" in report.errors[0].message
    assert session.exec(select(Credit)).first() is None


def test_missing_customer_reference_is_rejected(ctx, customer):
    report = _run(ctx, "credits", _csv("item_name,unit_price,due_date", f"Rice,100,{FUTURE}"))

    assert not report.ok
    assert "customer_code or customer_phone" in report.errors[0].message


def test_blocked_customer_is_rejected_before_the_write(ctx, session, customer):
    """CreditService would raise mid-batch; validation has to catch it first."""
    customer.status = CustomerStatus.BLOCKED
    session.add(customer)
    session.commit()

    report = _run(ctx, "credits", _csv(CREDIT_HEADER, f"CUST-0001,Rice,1,100,{FUTURE},,"))

    assert not report.ok
    assert "blocked" in report.errors[0].message.lower()
    assert session.exec(select(Credit)).first() is None


def test_due_date_before_issued_date_is_rejected(ctx, customer):
    report = _run(ctx, "credits", _csv(CREDIT_HEADER, f"CUST-0001,Rice,1,100,{PAST},{FUTURE},"))

    assert not report.ok
    assert report.errors[0].column == "due_date"


def test_ambiguous_dates_are_refused_rather_than_guessed(ctx, customer):
    """08/07/2026 is two different dates depending on where you live."""
    report = _run(ctx, "credits", _csv(CREDIT_HEADER, "CUST-0001,Rice,1,100,08/07/2026,,"))

    assert not report.ok
    assert report.errors[0].column == "due_date"
    assert "YYYY-MM-DD" in report.errors[0].message


def test_overpayment_is_caught_in_validation(ctx, session, customer):
    report = _run(ctx, "credits", _csv(CREDIT_HEADER, f"CUST-0001,Rice,1,100,{FUTURE},,500"))

    assert not report.ok
    assert report.errors[0].column == "initial_payment"
    assert session.exec(select(Credit)).first() is None


def test_credit_batch_is_atomic(ctx, session, customer):
    """Row 2 is valid, row 3 is not. Neither may land."""
    report = _run(
        ctx,
        "credits",
        _csv(
            CREDIT_HEADER,
            f"CUST-0001,Rice,1,100,{FUTURE},,",
            f"CUST-0001,Oil,1,notanumber,{FUTURE},,",
        ),
    )

    assert not report.ok
    assert report.created == 0
    assert session.exec(select(Credit)).first() is None
    session.refresh(customer)
    assert customer.credit_count == 0


# ---------------------------------------------------------------------------
# Tenancy
# ---------------------------------------------------------------------------
def test_cannot_reference_another_tenants_customer(ctx, session, business):
    """A code that exists -- for someone else -- must read as "not found"."""
    from app.models.business import Business

    other = Business(name="Rival Shop", slug="rival-shop", email="rival@x.bt")
    session.add(other)
    session.commit()
    session.add(Customer(business_id=other.id, code="CUST-0001", name="Their Customer"))
    session.commit()

    report = _run(ctx, "credits", _csv(CREDIT_HEADER, f"CUST-0001,Rice,1,100,{FUTURE},,"))

    assert not report.ok
    assert "No customer has the code CUST-0001" in report.errors[0].message


# ---------------------------------------------------------------------------
# Templates
# ---------------------------------------------------------------------------
def test_csv_template_is_headings_only(ctx):
    """No example rows -- see ImportService.template for why."""
    data, filename, content_type = ImportService(ctx).template("customers", "csv")

    text = data.decode("utf-8-sig")
    assert filename == "customers-import-template.csv"
    assert "csv" in content_type
    assert text.strip().splitlines() == [",".join(c.key for c in CUSTOMER_COLUMNS)]


def test_csv_template_carries_a_bom_for_excel(ctx):
    data, _, _ = ImportService(ctx).template("customers", "csv")
    assert data.startswith(b"\xef\xbb\xbf")


def test_xlsx_template_has_a_data_sheet_and_an_instructions_sheet(ctx):
    from openpyxl import load_workbook

    data, filename, _ = ImportService(ctx).template("credits", "xlsx")
    assert filename == "credits-import-template.xlsx"

    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Credits", "Instructions"]

    sheet = wb["Credits"]
    headings = [c.value for c in sheet[1]]
    assert headings[0] == "customer_code"
    assert "due_date" in headings
    assert sheet.max_row == 1  # headings only, no phantom sample row
    assert sheet.freeze_panes == "A2"


def test_a_downloaded_template_is_importable_once_filled_in(ctx, session):
    """The loop that matters: template -> type -> upload. It must round-trip."""
    from openpyxl import load_workbook

    data, _, _ = ImportService(ctx).template("customers", "xlsx")
    wb = load_workbook(io.BytesIO(data))
    sheet = wb["Customers"]
    headings = [c.value for c in sheet[1]]

    row = {"name": "Sonam Dorji", "phone": "17111111", "city": "Thimphu", "status": "ACTIVE"}
    for index, heading in enumerate(headings, start=1):
        if heading in row:
            sheet.cell(row=2, column=index, value=row[heading])

    filled = io.BytesIO()
    wb.save(filled)

    report = ImportService(ctx).run(
        ctx,
        dataset="customers",
        filename="customers-import-template.xlsx",
        data=filled.getvalue(),
        dry_run=False,
    )

    assert report.ok, report.errors
    assert report.created == 1
    saved = _customers(session)[0]
    assert saved.name == "Sonam Dorji"
    assert saved.city == "Thimphu"


def test_excel_native_types_do_not_leak_into_the_data(ctx, session, customer):
    """A date cell arrives as a datetime and a quantity as 2.0. Both must clean up."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.append(["customer_code", "item_name", "quantity", "unit_price", "due_date"])
    sheet.append(["CUST-0001", "Rice 5kg", 2.0, 450.0, TODAY + timedelta(days=30)])

    buf = io.BytesIO()
    wb.save(buf)

    report = ImportService(ctx).run(
        ctx, dataset="credits", filename="s.xlsx", data=buf.getvalue(), dry_run=False
    )

    assert report.ok, report.errors
    credit = session.exec(select(Credit)).first()
    assert credit.due_date == TODAY + timedelta(days=30)
    assert credit.grand_total == Decimal("900.00")


def test_unknown_dataset_is_rejected(ctx):
    with pytest.raises(ValidationError, match="Unknown import type"):
        _run(ctx, "invoices", _csv("a", "b"))
